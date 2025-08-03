from flask import render_template, request, redirect, url_for, session, flash, jsonify, make_response, send_file, Response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
from app import app, db
from models import Hotel, Event, EventForecast, MonthlyForecast, HotelActuals, User, Comment, Task, TaskComment, UserTaskFollow, HotelAssignment, ChatConversation, ChatMessage, EventSearch, ExternalEvent, EventSearchExport
from flask import send_file
from functools import wraps
import pandas as pd
from io import BytesIO, StringIO
import json
import csv
# Chart.js is used for frontend charts - no backend plotting library needed
import os
from datetime import datetime, timedelta, date
from sqlalchemy import or_, desc, and_
import logging
from chat_assistant import chat_assistant
from event_finder import EventFinderService

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Authentication Routes
@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration page"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')
        full_name = request.form.get('full_name', '').strip()
        
        # Validation
        if not username or not email or not password:
            flash('Please fill in all required fields.', 'error')
            return render_template('auth/register.html')
        
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return render_template('auth/register.html')
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('auth/register.html')
        
        # Check for existing users
        if User.query.filter_by(username=username).first():
            flash('Username already exists. Please choose a different one.', 'error')
            return render_template('auth/register.html')
        
        if User.query.filter_by(email=email).first():
            flash('Email already registered. Please use a different email or login.', 'error')
            return render_template('auth/register.html')
        
        # Create new user
        user = User(
            username=username,
            email=email,
            full_name=full_name or None
        )
        user.set_password(password)
        
        try:
            db.session.add(user)
            db.session.commit()
            flash('Registration successful! Please log in.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash('Registration failed. Please try again.', 'error')
            
    return render_template('auth/register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login page"""
    # If already logged in, redirect to dashboard
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        username_or_email = request.form.get('username_or_email', '').strip()
        password = request.form.get('password', '')
        remember_me = bool(request.form.get('remember_me'))
        
        print(f"DEBUG: Login attempt - username_or_email: {username_or_email}, password length: {len(password) if password else 0}")
        
        if not username_or_email or not password:
            flash('Please enter both username/email and password.', 'error')
            return render_template('auth/login.html')
        
        # Try to find user by username or email
        user = User.query.filter(
            (User.username == username_or_email) | (User.email == username_or_email)
        ).first()
        
        print(f"DEBUG: User found: {user is not None}")
        
        if user and user.check_password(password) and user.is_active:
            print(f"DEBUG: Password valid, logging in user: {user.username}")
            
            # Update last login time
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            # Log in the user
            login_user(user, remember=remember_me)
            flash(f'Welcome back, {user.get_display_name()}!', 'success')
            
            # Redirect to next page or dashboard
            next_page = request.args.get('next')
            print(f"DEBUG: Redirecting to dashboard after login")
            if next_page:
                return redirect(next_page)
            return redirect(url_for('dashboard'))
        else:
            print(f"DEBUG: Login failed - user exists: {user is not None}, password check: {user.check_password(password) if user else False}, active: {user.is_active if user else False}")
            flash('Invalid username/email or password.', 'error')
            
    return render_template('auth/login.html')

@app.route('/auth/logout')
@login_required
def auth_logout():
    """User logout"""
    logout_user()
    flash('You have been logged out successfully.', 'success')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard with Netflix-style carousels"""
    # Get user's recent forecasts (using username for compatibility with existing data)
    user_identifier = current_user.username
    recent_forecasts = db.session.query(EventForecast, Event, Hotel).join(
        Event, EventForecast.event_id == Event.id
    ).join(
        Hotel, EventForecast.hotel_id == Hotel.id
    ).filter(
        EventForecast.created_by == user_identifier
    ).order_by(EventForecast.updated_at.desc()).limit(5).all()
    
    # Get summary statistics
    total_forecasts = EventForecast.query.filter_by(created_by=user_identifier).count()
    total_hotels = db.session.query(EventForecast.hotel_id).filter_by(created_by=user_identifier).distinct().count()
    
    # Get hotels for carousel (prioritize assigned hotels first)
    assigned_hotel_ids = [assignment.hotel_id for assignment in current_user.hotel_assignments if assignment.is_active]
    assigned_hotels = Hotel.query.filter(Hotel.id.in_(assigned_hotel_ids)).all() if assigned_hotel_ids else []
    other_hotels = Hotel.query.filter(~Hotel.id.in_(assigned_hotel_ids)).limit(12).all() if assigned_hotel_ids else Hotel.query.limit(15).all()
    
    # Get events for carousel (recent and upcoming)
    upcoming_events = Event.query.filter(Event.end_date >= datetime.now().date()).order_by(Event.start_date).limit(15).all()
    
    # Get all users for assignment dropdown
    all_users = User.query.filter_by(is_active=True).order_by(User.full_name, User.username).all()
    
    return render_template('dashboard.html',
                         recent_forecasts=recent_forecasts,
                         total_forecasts=total_forecasts,
                         total_hotels=total_hotels,
                         assigned_hotels=assigned_hotels,
                         other_hotels=other_hotels,
                         upcoming_events=upcoming_events,
                         all_users=all_users,
                         date=date)

@app.route('/')
def index():
    # Redirect authenticated users to dashboard
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    # Get featured hotels for homepage
    featured_hotels = Hotel.query.limit(6).all()
    
    # Get recent activity stats
    recent_forecasts = EventForecast.query.filter(
        EventForecast.created_at >= datetime.now() - timedelta(days=30)
    ).count()
    
    total_hotels = Hotel.query.count()
    total_events = Event.query.count()
    
    return render_template('home.html',
                         featured_hotels=featured_hotels,
                         recent_forecasts=recent_forecasts,
                         total_hotels=total_hotels,
                         total_events=total_events)

@app.route('/hotel-selection', methods=['GET', 'POST'])
@login_required
def hotel_selection():
    # Handle pre-selected hotel from homepage or event page
    preselected_hotel = request.args.get('hotel')
    selected_event_id = request.args.get('event')
    
    if request.method == 'POST':
        hotel_code = request.form.get('hotel_code')
        hotel = Hotel.query.filter_by(hotel_code=hotel_code).first()
        
        if hotel:
            session['hotel_id'] = hotel.id
            session['hotel_code'] = hotel.hotel_code
            session['hotel_name'] = hotel.hotel_name
            session['city'] = hotel.city
            session['inventory'] = hotel.inventory
            session['user_name'] = current_user.get_display_name()
            
            # Store selected event if coming from event page
            if selected_event_id:
                session['selected_event_id'] = int(selected_event_id)
            
            return redirect(url_for('event_forecast'))
        else:
            flash('Hotel not found. Please select a valid hotel.', 'error')
    
    # Auto-select hotel if coming from homepage or event page
    if preselected_hotel:
        hotel = Hotel.query.filter_by(hotel_code=preselected_hotel).first()
        if hotel:
            session['hotel_id'] = hotel.id
            session['hotel_code'] = hotel.hotel_code
            session['hotel_name'] = hotel.hotel_name
            session['city'] = hotel.city
            session['inventory'] = hotel.inventory
            session['user_name'] = current_user.get_display_name()
            
            # Store selected event if coming from event page
            if selected_event_id:
                session['selected_event_id'] = int(selected_event_id)
            
            return redirect(url_for('event_forecast'))
    
    hotels = Hotel.query.all()
    return render_template('hotel_selection.html', hotels=hotels, preselected=preselected_hotel)

@app.route('/event-forecast')
@login_required
def event_forecast():
    if 'hotel_id' not in session:
        return redirect(url_for('hotel_selection'))
    
    city = session.get('city')
    events = Event.query.filter_by(city=city).all()
    
    # Get existing forecasts for this hotel
    hotel_id = session.get('hotel_id')
    forecasts = {}
    
    for event in events:
        event_forecasts = EventForecast.query.filter_by(event_id=event.id, hotel_id=hotel_id).all()
        forecasts[event.id] = {}
        for forecast in event_forecasts:
            # Fix: use forecast_date to calculate day_number instead of non-existent day_number field
            day_number = (forecast.forecast_date - event.start_date).days + 1
            forecasts[event.id][day_number] = {
                'revenue': forecast.revenue,
                'adr': forecast.adr,
                'occupancy': forecast.occupancy
            }
    
    # Get the hotel object for the template
    selected_hotel = Hotel.query.get(hotel_id)
    
    # Get selected event if coming from event page
    selected_event_id = session.pop('selected_event_id', None)
    selected_event = None
    if selected_event_id:
        selected_event = Event.query.get(selected_event_id)
    
    return render_template('event_forecast.html', 
                         events=events, 
                         forecasts=forecasts,
                         hotel_name=session.get('hotel_name'),
                         city=city,
                         selected_hotel=selected_hotel,
                         inventory=session.get('inventory'),
                         selected_event=selected_event)

# Monthly Forecast Routes
@app.route('/monthly-forecast/<int:hotel_id>')
@login_required
def monthly_forecast(hotel_id):
    """Display monthly forecast calendar for a hotel"""
    hotel = Hotel.query.get_or_404(hotel_id)
    current_month = datetime.now().strftime('%B %Y')
    
    return render_template('monthly_forecast.html', 
                         hotel=hotel,
                         current_month=current_month)

@app.route('/api/hotel-events/<int:hotel_id>')
@login_required
def get_hotel_events(hotel_id):
    """Get all events for a hotel's city"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    # Get all events in this hotel's city
    events = Event.query.filter_by(city=hotel.city).order_by(Event.start_date.desc()).all()
    
    events_data = []
    for event in events:
        events_data.append({
            'id': event.id,
            'name': event.event_name,
            'start_date': event.start_date.isoformat(),
            'end_date': event.end_date.isoformat(),
            'city': event.city
        })
    
    return jsonify({
        'success': True,
        'events': events_data
    })

@app.route('/api/monthly-forecast/<int:hotel_id>/<int:year>/<int:month>')
@login_required
def get_monthly_forecast_data(hotel_id, year, month):
    """Get monthly forecast data for calendar display"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    # Get start and end dates for the month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month + 1, 1) - timedelta(days=1)
    
    # Get monthly forecasts
    monthly_forecasts = MonthlyForecast.query.filter(
        MonthlyForecast.hotel_id == hotel_id,
        MonthlyForecast.forecast_date >= start_date,
        MonthlyForecast.forecast_date <= end_date
    ).all()
    
    # Get event forecasts for the same period
    event_forecasts = db.session.query(EventForecast, Event).join(
        Event, EventForecast.event_id == Event.id
    ).filter(
        EventForecast.hotel_id == hotel_id,
        EventForecast.forecast_date >= start_date,
        EventForecast.forecast_date <= end_date
    ).all()
    
    # Format data for calendar
    monthly_data = {}
    for forecast in monthly_forecasts:
        date_str = forecast.forecast_date.strftime('%Y-%m-%d')
        monthly_data[date_str] = {
            'revenue': float(forecast.revenue) if forecast.revenue else None,
            'adr': float(forecast.adr) if forecast.adr else None,
            'occupancy': float(forecast.occupancy) if forecast.occupancy else None,
            'room_nights': forecast.room_nights
        }
    
    event_forecast_data = {}
    for forecast, event in event_forecasts:
        date_str = forecast.forecast_date.strftime('%Y-%m-%d')
        if date_str not in event_forecast_data:
            event_forecast_data[date_str] = []
        event_forecast_data[date_str].append({
            'event_name': event.event_name,
            'revenue': float(forecast.revenue) if forecast.revenue else None,
            'adr': float(forecast.adr) if forecast.adr else None,
            'occupancy': float(forecast.occupancy) if forecast.occupancy else None
        })
    
    # Get all events in this hotel's city that overlap with the selected month
    hotel_events = Event.query.filter(
        Event.city == hotel.city,
        Event.start_date <= end_date,
        Event.end_date >= start_date
    ).all()
    
    # Map events to their date ranges
    events_by_date = {}
    for event in hotel_events:
        # Calculate date range for this event within the selected month
        event_start = max(event.start_date, start_date)
        event_end = min(event.end_date, end_date)
        
        current_date = event_start
        while current_date <= event_end:
            date_str = current_date.strftime('%Y-%m-%d')
            if date_str not in events_by_date:
                events_by_date[date_str] = []
            
            events_by_date[date_str].append({
                'id': event.id,
                'name': event.event_name,
                'start_date': event.start_date.isoformat(),
                'end_date': event.end_date.isoformat(),
                'website_url': event.website_url,
                'is_start_date': current_date == event.start_date,
                'is_end_date': current_date == event.end_date
            })
            current_date += timedelta(days=1)
    
    return jsonify({
        'monthly_forecasts': monthly_data,
        'event_forecasts': event_forecast_data,
        'events_by_date': events_by_date
    })

@app.route('/api/monthly-forecast/<int:hotel_id>', methods=['POST'])
@login_required
def save_monthly_forecast(hotel_id):
    """Save or update monthly forecast data"""
    hotel = Hotel.query.get_or_404(hotel_id)
    data = request.get_json()
    
    try:
        forecast_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        
        # Find or create forecast
        forecast = MonthlyForecast.query.filter_by(
            hotel_id=hotel_id,
            forecast_date=forecast_date
        ).first()
        
        if not forecast:
            forecast = MonthlyForecast(
                hotel_id=hotel_id,
                forecast_date=forecast_date,
                created_by=current_user.get_display_name()
            )
            db.session.add(forecast)
        
        # Update values
        forecast.revenue = data.get('revenue')
        forecast.adr = data.get('adr')
        forecast.occupancy = data.get('occupancy')
        
        # Calculate room nights
        if forecast.occupancy and hotel.inventory:
            forecast.room_nights = int((float(forecast.occupancy) / 100) * hotel.inventory)
        
        forecast.updated_at = datetime.utcnow()
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/download-monthly-template/<int:hotel_id>')
@login_required
def download_monthly_template(hotel_id):
    """Download Excel template for monthly forecast upload"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Monthly Forecast Template'
    
    # Headers
    headers = ['Date', 'Revenue', 'ADR', 'Occupancy %', 'Room Nights']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    
    # Add sample data for next 30 days
    start_date = date.today()
    for i in range(30):
        current_date = start_date + timedelta(days=i)
        ws.cell(row=i+2, column=1, value=current_date.strftime('%Y-%m-%d'))
        ws.cell(row=i+2, column=2, value='')  # Revenue
        ws.cell(row=i+2, column=3, value='')  # ADR
        ws.cell(row=i+2, column=4, value='')  # Occupancy
        ws.cell(row=i+2, column=5, value='')  # Room Nights
    
    # Adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    # Save to bytes
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'monthly_forecast_template_{hotel.hotel_code}_{date.today()}.xlsx'
    )

@app.route('/api/monthly-forecast/<int:hotel_id>/upload', methods=['POST'])
@login_required
def upload_monthly_forecast(hotel_id):
    """Upload monthly forecast data from Excel/CSV"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    try:
        # Read the file
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        count = 0
        created_by = current_user.get_display_name()
        
        for _, row in df.iterrows():
            # Parse date
            forecast_date = pd.to_datetime(row['Date']).date()
            
            # Find or create forecast
            forecast = MonthlyForecast.query.filter_by(
                hotel_id=hotel_id,
                forecast_date=forecast_date
            ).first()
            
            if not forecast:
                forecast = MonthlyForecast(
                    hotel_id=hotel_id,
                    forecast_date=forecast_date,
                    created_by=created_by
                )
                db.session.add(forecast)
            
            # Update values
            forecast.revenue = row.get('Revenue') if pd.notna(row.get('Revenue')) else None
            forecast.adr = row.get('ADR') if pd.notna(row.get('ADR')) else None
            forecast.occupancy = row.get('Occupancy %') if pd.notna(row.get('Occupancy %')) else None
            forecast.room_nights = row.get('Room Nights') if pd.notna(row.get('Room Nights')) else None
            forecast.updated_at = datetime.utcnow()
            
            count += 1
        
        db.session.commit()
        return jsonify({'success': True, 'count': count})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/hotel-forecast/<int:hotel_id>')
@login_required
def hotel_monthly_forecast(hotel_id):
    """Monthly forecast view for a specific hotel"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    # Get selected month (default to current month)
    month_str = request.args.get('month', datetime.now().strftime('%Y-%m'))
    try:
        selected_date = datetime.strptime(month_str, '%Y-%m')
    except ValueError:
        selected_date = datetime.now()
    
    # Calculate month boundaries
    month_start = selected_date.replace(day=1)
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(days=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1) - timedelta(days=1)
    
    # Get events for this hotel's city during the month
    events_in_month = Event.query.filter(
        Event.city == hotel.city,
        Event.start_date <= month_end.date(),
        Event.end_date >= month_start.date()
    ).all()
    
    # Get existing forecasts for this hotel and month
    existing_forecasts = {}
    for event in events_in_month:
        event_forecasts = EventForecast.query.filter_by(
            event_id=event.id, 
            hotel_id=hotel.id
        ).all()
        for forecast in event_forecasts:
            # Calculate actual date from event start + day_number
            forecast_date = event.start_date + timedelta(days=forecast.day_number - 1)
            if month_start.date() <= forecast_date <= month_end.date():
                existing_forecasts[forecast_date.strftime('%Y-%m-%d')] = {
                    'revenue': forecast.revenue,
                    'adr': forecast.adr,
                    'occupancy': forecast.occupancy,
                    'event': event
                }
    
    # Generate calendar days for the month
    calendar_days = []
    current_date = month_start
    while current_date <= month_end:
        date_str = current_date.strftime('%Y-%m-%d')
        
        # Check if this date has events
        events_on_date = []
        for event in events_in_month:
            if event.start_date <= current_date.date() <= event.end_date:
                events_on_date.append(event)
        
        calendar_days.append({
            'date': current_date.date(),
            'date_str': date_str,
            'day_of_week': current_date.strftime('%a'),
            'day_number': current_date.day,
            'events': events_on_date,
            'forecast': existing_forecasts.get(date_str),
            'is_weekend': current_date.weekday() >= 5
        })
        current_date += timedelta(days=1)
    
    # Calculate previous and next month dates for navigation
    if selected_date.month == 1:
        prev_month = selected_date.replace(year=selected_date.year - 1, month=12)
    else:
        prev_month = selected_date.replace(month=selected_date.month - 1)
    
    if selected_date.month == 12:
        next_month = selected_date.replace(year=selected_date.year + 1, month=1)
    else:
        next_month = selected_date.replace(month=selected_date.month + 1)
    
    return render_template('hotel_monthly_forecast.html',
                         hotel=hotel,
                         calendar_days=calendar_days,
                         selected_month=selected_date,
                         prev_month=prev_month,
                         next_month=next_month,
                         events_in_month=events_in_month,
                         month_name=selected_date.strftime('%B %Y'))

@app.route('/save-forecast', methods=['POST'])
@login_required
def save_forecast():
    """Save monthly forecast data for a specific hotel and date"""
    data = request.get_json()
    hotel_id = data.get('hotel_id')
    date_str = data.get('date')
    forecast_type = data.get('type')  # revenue, adr, occupancy
    value = data.get('value')
    
    try:
        forecast_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # For monthly forecasts, we'll create or update existing event-based forecasts
        # If no event exists for this date, we'll need to create a generic entry
        hotel = Hotel.query.get(hotel_id)
        if not hotel:
            return jsonify({'success': False, 'message': 'Hotel not found'})
        
        # Find events that cover this date
        events_on_date = Event.query.filter(
            Event.city == hotel.city,
            Event.start_date <= forecast_date,
            Event.end_date >= forecast_date
        ).all()
        
        created_by = current_user.get_display_name()
        
        if events_on_date:
            # Update existing event-based forecasts
            for event in events_on_date:
                # Calculate day number within the event
                day_number = (forecast_date - event.start_date).days + 1
                
                # Find or create forecast record
                forecast = EventForecast.query.filter_by(
                    event_id=event.id,
                    hotel_id=hotel_id,
                    day_number=day_number
                ).first()
                
                if not forecast:
                    forecast = EventForecast(
                        event_id=event.id,
                        hotel_id=hotel_id,
                        day_number=day_number,
                        created_by=created_by
                    )
                    db.session.add(forecast)
                
                # Update the specific forecast type
                if forecast_type == 'revenue':
                    forecast.revenue = value
                elif forecast_type == 'adr':
                    forecast.adr = value
                elif forecast_type == 'occupancy':
                    forecast.occupancy = value
                
                forecast.updated_at = datetime.now()
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/download-hotel-monthly-template/<int:hotel_id>')
@login_required
def download_hotel_monthly_template(hotel_id):
    """Download Excel template for monthly forecast"""
    hotel = Hotel.query.get_or_404(hotel_id)
    month_str = request.args.get('month', datetime.now().strftime('%Y-%m'))
    
    try:
        selected_date = datetime.strptime(month_str, '%Y-%m')
    except ValueError:
        selected_date = datetime.now()
    
    # Create Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Monthly Forecast Template"
    
    # Header row
    headers = ['Date', 'Day', 'Revenue', 'ADR', 'Occupancy %', 'Events']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Calculate month boundaries
    month_start = selected_date.replace(day=1)
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(days=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1) - timedelta(days=1)
    
    # Get events for this hotel's city during the month
    events_in_month = Event.query.filter(
        Event.city == hotel.city,
        Event.start_date <= month_end.date(),
        Event.end_date >= month_start.date()
    ).all()
    
    # Generate rows for each day
    current_date = month_start
    row = 2
    while current_date <= month_end:
        # Check for events on this date
        events_on_date = []
        for event in events_in_month:
            if event.start_date <= current_date.date() <= event.end_date:
                events_on_date.append(event.event_name)
        
        # Fill row data
        ws.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=current_date.strftime('%A'))
        ws.cell(row=row, column=3, value="")  # Revenue
        ws.cell(row=row, column=4, value="")  # ADR
        ws.cell(row=row, column=5, value="")  # Occupancy
        ws.cell(row=row, column=6, value=", ".join(events_on_date) if events_on_date else "")
        
        # Highlight weekends
        if current_date.weekday() >= 5:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Highlight event days
        if events_on_date:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        current_date += timedelta(days=1)
        row += 1
    
    # Auto-adjust column widths
    for col in range(1, 7):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Save to BytesIO
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'{hotel.hotel_name}_Monthly_Forecast_Template_{month_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/upload-hotel-monthly-forecast/<int:hotel_id>', methods=['POST'])
@login_required
def upload_hotel_monthly_forecast(hotel_id):
    """Upload Excel file with monthly forecast data"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Please upload an Excel file (.xlsx or .xls)'})
    
    try:
        import pandas as pd
        
        # Read Excel file
        df = pd.read_excel(file)
        
        # Validate required columns
        required_columns = ['Date', 'Revenue', 'ADR', 'Occupancy %']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'success': False, 'message': f'Missing columns: {", ".join(missing_columns)}'})
        
        created_by = current_user.get_display_name()
        uploaded_count = 0
        
        for _, row in df.iterrows():
            try:
                # Parse date
                if pd.isna(row['Date']):
                    continue
                    
                if isinstance(row['Date'], str):
                    forecast_date = datetime.strptime(row['Date'], '%Y-%m-%d').date()
                else:
                    forecast_date = row['Date'].date() if hasattr(row['Date'], 'date') else row['Date']
                
                # Get values
                revenue = float(row['Revenue']) if pd.notna(row['Revenue']) and row['Revenue'] != '' else None
                adr = float(row['ADR']) if pd.notna(row['ADR']) and row['ADR'] != '' else None
                occupancy = float(row['Occupancy %']) if pd.notna(row['Occupancy %']) and row['Occupancy %'] != '' else None
                
                # Skip if all values are empty
                if revenue is None and adr is None and occupancy is None:
                    continue
                
                # Find events that cover this date
                events_on_date = Event.query.filter(
                    Event.city == hotel.city,
                    Event.start_date <= forecast_date,
                    Event.end_date >= forecast_date
                ).all()
                
                if events_on_date:
                    # Update existing event-based forecasts
                    for event in events_on_date:
                        # Calculate day number within the event
                        day_number = (forecast_date - event.start_date).days + 1
                        
                        # Find or create forecast record
                        forecast = EventForecast.query.filter_by(
                            event_id=event.id,
                            hotel_id=hotel_id,
                            day_number=day_number
                        ).first()
                        
                        if not forecast:
                            forecast = EventForecast(
                                event_id=event.id,
                                hotel_id=hotel_id,
                                day_number=day_number,
                                created_by=created_by
                            )
                            db.session.add(forecast)
                        
                        # Update values
                        if revenue is not None:
                            forecast.revenue = revenue
                        if adr is not None:
                            forecast.adr = adr
                        if occupancy is not None:
                            forecast.occupancy = occupancy
                        
                        forecast.updated_at = datetime.now()
                        uploaded_count += 1
                
            except Exception as e:
                print(f"Error processing row: {e}")
                continue
        
        db.session.commit()
        return jsonify({
            'success': True, 
            'message': f'Successfully uploaded {uploaded_count} forecast entries'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error processing file: {str(e)}'})

@app.route('/get-monthly-actuals/<int:hotel_id>')
@login_required
def get_monthly_actuals(hotel_id):
    """Get actual data for comparison with monthly forecasts"""
    hotel = Hotel.query.get_or_404(hotel_id)
    month_str = request.args.get('month', datetime.now().strftime('%Y-%m'))
    
    try:
        selected_date = datetime.strptime(month_str, '%Y-%m')
    except ValueError:
        selected_date = datetime.now()
    
    # Calculate month boundaries
    month_start = selected_date.replace(day=1)
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(days=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1) - timedelta(days=1)
    
    # Get actual data
    actuals = HotelActuals.query.filter(
        HotelActuals.hotel_id == hotel_id,
        HotelActuals.date >= month_start.date(),
        HotelActuals.date <= month_end.date()
    ).all()
    
    # Get forecast data for comparison
    events_in_month = Event.query.filter(
        Event.city == hotel.city,
        Event.start_date <= month_end.date(),
        Event.end_date >= month_start.date()
    ).all()
    
    forecast_data = {}
    for event in events_in_month:
        event_forecasts = EventForecast.query.filter_by(
            event_id=event.id,
            hotel_id=hotel_id
        ).all()
        for forecast in event_forecasts:
            forecast_date = event.start_date + timedelta(days=forecast.day_number - 1)
            if month_start.date() <= forecast_date <= month_end.date():
                forecast_data[forecast_date.strftime('%Y-%m-%d')] = {
                    'revenue': forecast.revenue,
                    'adr': forecast.adr,
                    'occupancy': forecast.occupancy
                }
    
    # Create comparison data
    comparison_data = []
    for actual in actuals:
        date_str = actual.date.strftime('%Y-%m-%d')
        forecast = forecast_data.get(date_str, {})
        
        actual_occupancy = (actual.ty_room_nights / hotel.inventory * 100) if hotel.inventory and actual.ty_room_nights else 0
        
        comparison_data.append({
            'date': date_str,
            'actual_revenue': actual.ty_revenue or 0,
            'forecast_revenue': forecast.get('revenue', 0) or 0,
            'actual_adr': actual.ty_adr or 0,
            'forecast_adr': forecast.get('adr', 0) or 0,
            'actual_occupancy': actual_occupancy,
            'forecast_occupancy': forecast.get('occupancy', 0) or 0
        })
    
    return jsonify({'success': True, 'data': comparison_data})

@app.route('/validate-forecast-data', methods=['POST'])
@login_required
def validate_forecast_data():
    """Validate forecast data for accuracy and consistency"""
    data = request.get_json()
    hotel_id = data.get('hotel_id')
    month_str = data.get('month')
    
    hotel = Hotel.query.get_or_404(hotel_id)
    
    try:
        selected_date = datetime.strptime(month_str, '%Y-%m')
    except ValueError:
        return jsonify({'success': False, 'message': 'Invalid month format'})
    
    # Get forecast data for validation
    month_start = selected_date.replace(day=1)
    if month_start.month == 12:
        month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(days=1)
    else:
        month_end = month_start.replace(month=month_start.month + 1) - timedelta(days=1)
    
    events_in_month = Event.query.filter(
        Event.city == hotel.city,
        Event.start_date <= month_end.date(),
        Event.end_date >= month_start.date()
    ).all()
    
    validation_results = []
    errors = []
    warnings = []
    
    for event in events_in_month:
        forecasts = EventForecast.query.filter_by(
            event_id=event.id,
            hotel_id=hotel_id
        ).all()
        
        for forecast in forecasts:
            forecast_date = event.start_date + timedelta(days=forecast.day_number - 1)
            if month_start.date() <= forecast_date <= month_end.date():
                # Validate revenue and ADR consistency
                if forecast.revenue and forecast.adr and forecast.occupancy:
                    calculated_room_nights = (forecast.occupancy / 100) * hotel.inventory
                    expected_revenue = calculated_room_nights * forecast.adr
                    revenue_variance = abs(forecast.revenue - expected_revenue) / expected_revenue * 100
                    
                    if revenue_variance > 10:  # More than 10% variance
                        errors.append({
                            'date': forecast_date.strftime('%Y-%m-%d'),
                            'type': 'Revenue Inconsistency',
                            'message': f'Revenue ({forecast.revenue:,.0f}) doesn\'t match ADR × Occupancy calculation ({expected_revenue:,.0f})',
                            'severity': 'high'
                        })
                
                # Check for unrealistic values
                if forecast.occupancy and forecast.occupancy > 100:
                    errors.append({
                        'date': forecast_date.strftime('%Y-%m-%d'),
                        'type': 'Invalid Occupancy',
                        'message': f'Occupancy cannot exceed 100% (current: {forecast.occupancy}%)',
                        'severity': 'high'
                    })
                
                if forecast.adr and forecast.adr < 50:
                    warnings.append({
                        'date': forecast_date.strftime('%Y-%m-%d'),
                        'type': 'Low ADR',
                        'message': f'ADR seems unusually low: ${forecast.adr:.2f}',
                        'severity': 'medium'
                    })
                
                # Check for missing data on event days
                if not forecast.revenue or not forecast.adr or not forecast.occupancy:
                    warnings.append({
                        'date': forecast_date.strftime('%Y-%m-%d'),
                        'type': 'Incomplete Data',
                        'message': 'Missing forecast data for event day',
                        'severity': 'medium'
                    })
    
    return jsonify({
        'success': True,
        'validation_results': {
            'errors': errors,
            'warnings': warnings,
            'total_issues': len(errors) + len(warnings)
        }
    })

@app.route('/generate-performance-recommendations/<int:hotel_id>')
@login_required  
def generate_performance_recommendations(hotel_id):
    """Generate personalized performance recommendations"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    # Get recent performance data - use 1 year for more data availability
    one_year_ago = datetime.now() - timedelta(days=365)
    
    actuals = HotelActuals.query.filter(
        HotelActuals.hotel_id == hotel_id,
        HotelActuals.date >= one_year_ago.date()
    ).all()
    
    forecasts = db.session.query(EventForecast).join(Event).filter(
        EventForecast.hotel_id == hotel_id,
        Event.start_date >= one_year_ago.date()
    ).all()
    
    recommendations = []
    
    # Analyze occupancy trends
    if actuals:
        occupancy_records = [actual for actual in actuals if actual.ty_room_nights]
        if occupancy_records:
            avg_occupancy = sum((actual.ty_room_nights / hotel.inventory * 100) for actual in occupancy_records) / len(occupancy_records)
            
            if avg_occupancy < 70:
                recommendations.append({
                    'category': 'Occupancy',
                    'priority': 'high',
                    'title': 'Improve Occupancy Rate',
                    'description': f'Current average occupancy is {avg_occupancy:.1f}%. Consider revenue management strategies to increase bookings.',
                    'actions': [
                        'Review pricing strategy for low-demand periods',
                        'Enhance marketing efforts during slow seasons',
                        'Partner with booking platforms for better visibility'
                    ]
                })
            elif avg_occupancy > 90:
                recommendations.append({
                    'category': 'Occupancy',
                    'priority': 'medium',
                    'title': 'High Occupancy Optimization',
                    'description': f'Excellent occupancy at {avg_occupancy:.1f}%. Focus on rate optimization to maximize revenue.',
                    'actions': [
                        'Implement dynamic pricing during peak periods',
                        'Consider rate increases for high-demand dates',
                        'Optimize room type allocation'
                    ]
                })
        
        # Analyze ADR performance
        adr_records = [actual for actual in actuals if actual.ty_adr]
        if adr_records:
            avg_adr = sum(float(actual.ty_adr) for actual in adr_records) / len(adr_records)
            
            # Compare with forecasts
            forecast_adrs = [float(f.adr) for f in forecasts if f.adr]
            if forecast_adrs:
                avg_forecast_adr = sum(forecast_adrs) / len(forecast_adrs)
                adr_variance = (avg_adr - avg_forecast_adr) / avg_forecast_adr * 100
                
                if adr_variance < -10:
                    recommendations.append({
                        'category': 'ADR',
                        'priority': 'medium',
                        'title': 'ADR Below Forecast',
                        'description': f'Actual ADR is {abs(adr_variance):.1f}% below forecast. Focus on rate optimization.',
                        'actions': [
                            'Review competitive pricing in your market',
                            'Implement dynamic pricing strategies',
                            'Focus on value-added packages'
                        ]
                    })
                elif adr_variance > 15:
                    recommendations.append({
                        'category': 'ADR',
                        'priority': 'low',
                        'title': 'ADR Exceeding Forecast',
                        'description': f'Actual ADR is {adr_variance:.1f}% above forecast. Great performance!',
                        'actions': [
                            'Maintain current pricing strategy',
                            'Explore premium service offerings',
                            'Consider upselling opportunities'
                        ]
                    })
            
            # General ADR analysis
            if avg_adr < 150:
                recommendations.append({
                    'category': 'ADR',
                    'priority': 'medium',
                    'title': 'ADR Growth Opportunity',
                    'description': f'Current ADR is ${avg_adr:.2f}. There may be room for rate optimization.',
                    'actions': [
                        'Analyze competitor pricing in your market',
                        'Review amenities and service quality',
                        'Consider premium room categories'
                    ]
                })
    
    # Analyze forecast accuracy
    forecast_accuracy_issues = 0
    for forecast in forecasts:
        # Find corresponding actual data  
        if forecast.forecast_date:
            actual = HotelActuals.query.filter_by(
                hotel_id=hotel_id,
                date=forecast.forecast_date
            ).first()
            
            if actual and forecast.revenue and actual.ty_revenue:
                variance = abs(forecast.revenue - actual.ty_revenue) / actual.ty_revenue * 100
                if variance > 20:
                    forecast_accuracy_issues += 1
    
    if forecast_accuracy_issues > 5:
        recommendations.append({
            'category': 'Forecasting',
            'priority': 'medium',
            'title': 'Improve Forecast Accuracy',
            'description': f'Multiple forecasts show significant variance from actual results.',
            'actions': [
                'Review historical data patterns more carefully',
                'Consider market factors in forecasting',
                'Update forecasting methodology'
            ]
        })
    
    # Revenue optimization recommendations
    revenue_records = [actual for actual in actuals if actual.ty_revenue]
    if revenue_records:
        total_revenue = sum(float(actual.ty_revenue) for actual in revenue_records)
        avg_daily_revenue = total_revenue / len(revenue_records)
        
        recommendations.append({
            'category': 'Revenue',
            'priority': 'low',
            'title': 'Revenue Performance Analysis',
            'description': f'Average daily revenue: ${avg_daily_revenue:,.0f}. Consider these optimization strategies.',
            'actions': [
                'Analyze peak demand periods for rate increases',
                'Implement upselling programs for amenities',
                'Review group booking strategies',
                'Optimize length-of-stay pricing'
            ]
        })
    
    # If no data available, provide general recommendations
    if not actuals:
        recommendations.append({
            'category': 'Data',
            'priority': 'high',
            'title': 'Limited Performance Data',
            'description': 'No recent performance data available for detailed analysis.',
            'actions': [
                'Upload recent actual performance data',
                'Review data collection processes',
                'Ensure regular reporting is in place'
            ]
        })
        
        # Add general best practices
        recommendations.append({
            'category': 'General',
            'priority': 'medium',
            'title': 'Revenue Management Best Practices',
            'description': 'Implement these strategies to optimize hotel performance.',
            'actions': [
                'Establish dynamic pricing based on demand',
                'Focus on direct booking channels',
                'Implement revenue management systems',
                'Regular competitive market analysis'
            ]
        })
    
    return jsonify({
        'success': True,
        'recommendations': recommendations,
        'hotel_name': hotel.hotel_name
    })

@app.route('/hotel-rankings')
@login_required
def hotel_rankings():
    """Display hotel rankings dashboard"""
    return render_template('hotel_rankings.html')

@app.route('/get-hotel-rankings-data')
@login_required
def get_hotel_rankings_data():
    """Get hotel ranking data based on filter criteria"""
    ranking_type = request.args.get('type', 'actual')  # actual, forecast, impact
    metric = request.args.get('metric', 'revenue')  # revenue, adr, occupancy
    period = request.args.get('period', '365')  # days - default to last year
    
    if period == 'all':
        # Use earliest possible date for all-time data
        start_date = datetime(2020, 1, 1)  # Reasonable start date for hotel data
    else:
        days_back = int(period)
        start_date = datetime.now() - timedelta(days=days_back)
    
    hotels = Hotel.query.all()
    ranking_data = []
    
    for hotel in hotels:
        hotel_data = {
            'hotel_id': hotel.id,
            'hotel_name': hotel.hotel_name,
            'city': hotel.city,
            'inventory': hotel.inventory,
            'data_points': 0  # Track number of data points used
        }
        
        if ranking_type == 'actual':
            # Get actual performance data
            actuals = HotelActuals.query.filter(
                HotelActuals.hotel_id == hotel.id,
                HotelActuals.date >= start_date.date()
            ).all()
            
            if actuals:
                hotel_data['data_points'] = len(actuals)
                if metric == 'revenue':
                    hotel_data['value'] = sum(a.ty_revenue or 0 for a in actuals)
                    hotel_data['avg_value'] = hotel_data['value'] / len(actuals)
                elif metric == 'adr':
                    adr_values = [float(a.ty_adr) for a in actuals if a.ty_adr]
                    hotel_data['value'] = sum(adr_values) / len(adr_values) if adr_values else 0
                    hotel_data['avg_value'] = hotel_data['value']
                    hotel_data['data_points'] = len(adr_values)
                elif metric == 'occupancy':
                    occ_values = [(a.ty_room_nights / hotel.inventory * 100) for a in actuals if a.ty_room_nights and hotel.inventory]
                    hotel_data['value'] = sum(occ_values) / len(occ_values) if occ_values else 0
                    hotel_data['avg_value'] = hotel_data['value']
                    hotel_data['data_points'] = len(occ_values)
            else:
                hotel_data['value'] = 0
                hotel_data['avg_value'] = 0
                
        elif ranking_type == 'forecast':
            # Get forecast performance data
            forecasts = db.session.query(EventForecast).join(Event).filter(
                EventForecast.hotel_id == hotel.id,
                Event.start_date >= start_date.date()
            ).all()
            
            if forecasts:
                if metric == 'revenue':
                    rev_values = [f.revenue for f in forecasts if f.revenue]
                    hotel_data['value'] = sum(rev_values) if rev_values else 0
                    hotel_data['avg_value'] = hotel_data['value'] / len(rev_values) if rev_values else 0
                    hotel_data['data_points'] = len(rev_values)
                elif metric == 'adr':
                    adr_values = [float(f.adr) for f in forecasts if f.adr]
                    hotel_data['value'] = sum(adr_values) / len(adr_values) if adr_values else 0
                    hotel_data['avg_value'] = hotel_data['value']
                    hotel_data['data_points'] = len(adr_values)
                elif metric == 'occupancy':
                    occ_values = [f.occupancy for f in forecasts if f.occupancy]
                    hotel_data['value'] = sum(occ_values) / len(occ_values) if occ_values else 0
                    hotel_data['avg_value'] = hotel_data['value']
                    hotel_data['data_points'] = len(occ_values)
            else:
                hotel_data['value'] = 0
                hotel_data['avg_value'] = 0
                
        elif ranking_type == 'impact':
            # Calculate impact (actual vs forecast variance)
            actuals = HotelActuals.query.filter(
                HotelActuals.hotel_id == hotel.id,
                HotelActuals.date >= start_date.date()
            ).all()
            
            forecasts = db.session.query(EventForecast).join(Event).filter(
                EventForecast.hotel_id == hotel.id,
                Event.start_date >= start_date.date()
            ).all()
            
            # Match actuals with forecasts by date
            total_impact = 0
            comparison_count = 0
            
            for actual in actuals:
                for forecast in forecasts:
                    forecast_event = Event.query.get(forecast.event_id)
                    if forecast_event:
                        forecast_date = forecast_event.start_date + timedelta(days=forecast.day_number - 1)
                        if forecast_date == actual.date:
                            if metric == 'revenue' and actual.ty_revenue and forecast.revenue:
                                impact = actual.ty_revenue - forecast.revenue
                                total_impact += impact
                                comparison_count += 1
                            elif metric == 'adr' and actual.ty_adr and forecast.adr:
                                impact = float(actual.ty_adr) - float(forecast.adr)
                                total_impact += impact
                                comparison_count += 1
                            elif metric == 'occupancy' and actual.ty_room_nights and forecast.occupancy and hotel.inventory:
                                actual_occ = (actual.ty_room_nights / hotel.inventory * 100)
                                impact = actual_occ - forecast.occupancy
                                total_impact += impact
                                comparison_count += 1
            
            hotel_data['value'] = total_impact
            hotel_data['avg_value'] = total_impact / comparison_count if comparison_count > 0 else 0
            hotel_data['data_points'] = comparison_count
        
        ranking_data.append(hotel_data)
    
    # Sort by value (descending)
    ranking_data.sort(key=lambda x: x['value'], reverse=True)
    
    # Add ranking position
    for i, hotel_data in enumerate(ranking_data):
        hotel_data['rank'] = i + 1
    
    return jsonify({
        'success': True,
        'rankings': ranking_data,
        'criteria': {
            'type': ranking_type,
            'metric': metric,
            'period': period
        }
    })

@app.route('/get-hotel-ranking-chart/<int:hotel_id>')
@login_required
def get_hotel_ranking_chart(hotel_id):
    """Get chart data for a specific hotel in rankings"""
    ranking_type = request.args.get('type', 'actual')
    metric = request.args.get('metric', 'revenue')
    period = request.args.get('period', '365')
    
    hotel = Hotel.query.get_or_404(hotel_id)
    
    if period == 'all':
        # Use earliest possible date for all-time data
        start_date = datetime(2020, 1, 1)  # Reasonable start date for hotel data
    else:
        days_back = int(period)
        start_date = datetime.now() - timedelta(days=days_back)
    
    chart_data = []
    
    if ranking_type == 'actual':
        actuals = HotelActuals.query.filter(
            HotelActuals.hotel_id == hotel_id,
            HotelActuals.date >= start_date.date()
        ).order_by(HotelActuals.date).all()
        
        for actual in actuals:
            if metric == 'revenue':
                value = float(actual.ty_revenue) if actual.ty_revenue else 0
            elif metric == 'adr':
                value = float(actual.ty_adr) if actual.ty_adr else 0
            elif metric == 'occupancy':
                value = (actual.ty_room_nights / hotel.inventory * 100) if actual.ty_room_nights and hotel.inventory else 0
            
            chart_data.append({
                'date': actual.date.strftime('%Y-%m-%d'),
                'value': value,
                'label': actual.date.strftime('%b %d')
            })
    
    # Add more chart data logic for forecast and impact types...
    
    return jsonify({
        'success': True,
        'chart_data': chart_data,
        'hotel_name': hotel.hotel_name
    })

@app.route('/save-event-forecast', methods=['POST'])
def save_event_forecast():
    if 'hotel_id' not in session:
        return jsonify({'success': False, 'message': 'Session expired'})
    
    data = request.get_json()
    event_id = data.get('event_id')
    day_number = data.get('day_number')
    forecast_type = data.get('type')  # revenue, adr, occupancy
    value = data.get('value')
    
    hotel_id = session.get('hotel_id')
    created_by = session.get('user_name', 'Anonymous User')
    
    # Convert day_number to forecast_date
    event = Event.query.get(event_id)
    if not event:
        return jsonify({'success': False, 'message': 'Event not found'})
    
    forecast_date = event.start_date + timedelta(days=int(day_number) - 1)
    
    # Find or create forecast record
    forecast = EventForecast.query.filter_by(
        event_id=event_id, 
        hotel_id=hotel_id, 
        forecast_date=forecast_date
    ).first()
    
    if not forecast:
        forecast = EventForecast(
            event_id=event_id,
            hotel_id=hotel_id,
            forecast_date=forecast_date,
            created_by=created_by
        )
        db.session.add(forecast)
    else:
        forecast.created_by = created_by
        forecast.updated_at = datetime.utcnow()
    
    # Update the specific field
    if forecast_type == 'revenue':
        forecast.revenue = float(value) if value else None
    elif forecast_type == 'adr':
        forecast.adr = float(value) if value else None
    elif forecast_type == 'occupancy':
        forecast.occupancy = float(value) if value else None
    
    try:
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/export-csv')
def export_csv():
    if 'hotel_id' not in session:
        return redirect(url_for('hotel_selection'))
    
    hotel_id = session.get('hotel_id')
    city = session.get('city')
    
    # Get all events for the city
    events = Event.query.filter_by(city=city).all()
    
    # Create export data
    export_data = []
    
    for event in events:
        forecasts = EventForecast.query.filter_by(event_id=event.id, hotel_id=hotel_id).all()
        # Create a dict mapping forecast_date to forecast object for easier lookup
        forecast_dict = {f.forecast_date: f for f in forecasts}
        
        for day in range(1, 16):  # Days 1-15
            forecast_date = event.start_date + timedelta(days=day - 1)
            forecast = forecast_dict.get(forecast_date)
            
            row = {
                'Hotel Code': session.get('hotel_code'),
                'Hotel Name': session.get('hotel_name'),
                'City': city,
                'Event Name': event.event_name,
                'Start Date': event.start_date.strftime('%Y-%m-%d'),
                'End Date': event.end_date.strftime('%Y-%m-%d'),
                'Duration': event.duration,
                'Day': f'D{day}',
                'Revenue': forecast.revenue if forecast and forecast.revenue else '',
                'ADR': forecast.adr if forecast and forecast.adr else '',
                'Occupancy': forecast.occupancy if forecast and forecast.occupancy else ''
            }
            export_data.append(row)
    
    # Create DataFrame and CSV
    df = pd.DataFrame(export_data)
    
    # Create response
    output = BytesIO()
    df.to_csv(output, index=False)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=forecast_{session.get("hotel_code")}_{city}.csv'
    
    return response

@app.route('/reset-forecast', methods=['POST'])
def reset_forecast():
    if 'hotel_id' not in session:
        return jsonify({'success': False, 'message': 'Session expired'})
    
    hotel_id = session.get('hotel_id')
    city = session.get('city')
    
    # Delete all forecasts for this hotel and city
    events = Event.query.filter_by(city=city).all()
    event_ids = [event.id for event in events]
    
    EventForecast.query.filter(
        EventForecast.event_id.in_(event_ids),
        EventForecast.hotel_id == hotel_id
    ).delete(synchronize_session=False)
    
    try:
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/change-hotel')
def change_hotel():
    session.clear()
    return redirect(url_for('hotel_selection'))

# Management Routes
@app.route('/manage-hotels')
@login_required
def manage_hotels():
    search_query = request.args.get('search', '').strip()
    if search_query:
        hotels = Hotel.query.filter(
            or_(Hotel.hotel_code.contains(search_query),
                Hotel.hotel_name.contains(search_query),
                Hotel.city.contains(search_query))
        ).all()
    else:
        hotels = Hotel.query.all()
    return render_template('manage_hotels.html', hotels=hotels, search_query=search_query)

@app.route('/add-hotel', methods=['GET', 'POST'])
def add_hotel():
    if request.method == 'POST':
        hotel_code = request.form.get('hotel_code').strip().upper()
        hotel_name = request.form.get('hotel_name').strip()
        city = request.form.get('city').strip()
        inventory = request.form.get('inventory', type=int)
        
        # Check if hotel code already exists
        existing_hotel = Hotel.query.filter_by(hotel_code=hotel_code).first()
        if existing_hotel:
            flash('Hotel code already exists. Please use a different code.', 'error')
            return render_template('add_hotel.html')
        
        hotel = Hotel(
            hotel_code=hotel_code,
            hotel_name=hotel_name,
            city=city,
            inventory=inventory,
            hotel_link=request.form.get('hotel_link') if request.form.get('hotel_link') else None,
            address_link=request.form.get('address_link') if request.form.get('address_link') else None
        )
        
        try:
            db.session.add(hotel)
            db.session.commit()
            flash(f'Hotel {hotel_code} added successfully!', 'success')
            return redirect(url_for('manage_hotels'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding hotel: {str(e)}', 'error')
    
    return render_template('add_hotel.html')

@app.route('/hotel/<int:hotel_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_hotel(hotel_id):
    """Edit hotel information including image upload"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    if request.method == 'POST':
        # Update hotel information
        hotel.hotel_name = request.form.get('hotel_name', '').strip()
        hotel.city = request.form.get('city', '').strip()
        hotel.inventory = int(request.form.get('inventory', 0))
        hotel.hotel_link = request.form.get('hotel_link', '').strip() or None
        hotel.address_link = request.form.get('address_link', '').strip() or None
        hotel.description = request.form.get('description', '').strip() or None
        
        # Handle image URL
        image_url = request.form.get('image_url', '').strip()
        if image_url:
            hotel.image_url = image_url
        
        try:
            db.session.commit()
            flash('Hotel updated successfully!', 'success')
            return redirect(url_for('hotel_detail', hotel_id=hotel.id))
        except Exception as e:
            db.session.rollback()
            flash('Error updating hotel. Please try again.', 'error')
    
    return render_template('edit_hotel.html', hotel=hotel)

@app.route('/delete-hotel/<int:hotel_id>', methods=['POST'])
def delete_hotel(hotel_id):
    hotel = Hotel.query.get_or_404(hotel_id)
    
    # Check if hotel has associated forecasts
    forecast_count = EventForecast.query.filter_by(hotel_id=hotel_id).count()
    if forecast_count > 0:
        flash(f'Cannot delete hotel {hotel.hotel_code}. It has {forecast_count} associated forecasts.', 'error')
        return redirect(url_for('manage_hotels'))
    
    try:
        db.session.delete(hotel)
        db.session.commit()
        flash(f'Hotel {hotel.hotel_code} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting hotel: {str(e)}', 'error')
    
    return redirect(url_for('manage_hotels'))

@app.route('/manage-events')
@login_required
def manage_events():
    search_query = request.args.get('search', '').strip()
    if search_query:
        events = Event.query.filter(
            or_(Event.event_name.contains(search_query),
                Event.city.contains(search_query))
        ).all()
    else:
        events = Event.query.all()
    return render_template('manage_events.html', events=events, search_query=search_query)

@app.route('/add-event', methods=['GET', 'POST'])
def add_event():
    if request.method == 'POST':
        event_name = request.form.get('event_name').strip()
        start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        city = request.form.get('city').strip()
        
        if start_date > end_date:
            flash('Start date cannot be after end date.', 'error')
            return render_template('add_event.html')
        
        event = Event(
            event_name=event_name,
            start_date=start_date,
            end_date=end_date,
            city=city
        )
        
        try:
            db.session.add(event)
            db.session.commit()
            flash(f'Event "{event_name}" added successfully!', 'success')
            return redirect(url_for('manage_events'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding event: {str(e)}', 'error')
    
    return render_template('add_event.html')

@app.route('/event/<int:event_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_event(event_id):
    """Edit event information including image upload"""
    event = Event.query.get_or_404(event_id)
    
    if request.method == 'POST':
        # Update event information
        event.event_name = request.form.get('event_name', '').strip()
        event.city = request.form.get('city', '').strip()
        event.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        event.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        event.website_url = request.form.get('website_url', '').strip() or None
        event.description = request.form.get('description', '').strip() or None
        
        # Handle image URL
        image_url = request.form.get('image_url', '').strip()
        if image_url:
            event.image_url = image_url
        
        try:
            db.session.commit()
            flash('Event updated successfully!', 'success')
            return redirect(url_for('event_detail', event_id=event.id))
        except Exception as e:
            db.session.rollback()
            flash('Error updating event. Please try again.', 'error')
    
    return render_template('edit_event.html', event=event)

@app.route('/delete-event/<int:event_id>', methods=['POST'])
def delete_event(event_id):
    event = Event.query.get_or_404(event_id)
    
    # Check if event has associated forecasts
    forecast_count = EventForecast.query.filter_by(event_id=event_id).count()
    if forecast_count > 0:
        flash(f'Cannot delete event "{event.event_name}". It has {forecast_count} associated forecasts.', 'error')
        return redirect(url_for('manage_events'))
    
    try:
        db.session.delete(event)
        db.session.commit()
        flash(f'Event "{event.event_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting event: {str(e)}', 'error')
    
    return redirect(url_for('manage_events'))

@app.route('/search-property')
def search_property():
    search_query = request.args.get('search', '').strip()
    hotels = []
    events = []
    forecasts = []
    
    if search_query:
        # Search hotels
        hotels = Hotel.query.filter(
            or_(Hotel.hotel_code.contains(search_query),
                Hotel.hotel_name.contains(search_query),
                Hotel.city.contains(search_query))
        ).all()
        
        # If we found hotels, get their events and forecasts
        if hotels:
            hotel_ids = [h.id for h in hotels]
            cities = list(set([h.city for h in hotels]))
            
            # Get events in the same cities
            events = Event.query.filter(Event.city.in_(cities)).all()
            
            # Get forecasts for these hotels
            forecasts = db.session.query(EventForecast, Event, Hotel).join(
                Event, EventForecast.event_id == Event.id
            ).join(
                Hotel, EventForecast.hotel_id == Hotel.id
            ).filter(
                EventForecast.hotel_id.in_(hotel_ids)
            ).all()
    
    return render_template('search_property.html', 
                         hotels=hotels, 
                         events=events, 
                         forecasts=forecasts,
                         search_query=search_query)

@app.route('/set-user', methods=['POST'])
def set_user():
    user_name = request.form.get('user_name', '').strip()
    if user_name:
        session['user_name'] = user_name
        
        # Find or create user profile
        user = User.query.filter_by(username=user_name).first()
        if not user:
            user = User(username=user_name)
            db.session.add(user)
            db.session.commit()
        
        # Update last login
        user.last_login = datetime.utcnow()
        db.session.commit()
        
        flash(f'Welcome, {user_name}!', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/logout')
def logout():
    """Log out the current user"""
    if 'user_name' in session:
        user_name = session['user_name']
        session.clear()
        flash(f'Goodbye, {user_name}! You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route('/edit-profile', methods=['GET', 'POST'])
@login_required
def edit_profile():
    """Edit user profile"""
    user = current_user
    
    if request.method == 'POST':
        # Update user information
        user.full_name = request.form.get('full_name', '').strip() or None
        user.email = request.form.get('email', '').strip() or None
        user.department = request.form.get('department', '').strip() or None
        user.role = request.form.get('role', '').strip() or None
        user.phone = request.form.get('phone', '').strip() or None
        user.bio = request.form.get('bio', '').strip() or None
        user.theme_preference = request.form.get('theme_preference', 'auto')
        user.email_notifications = bool(request.form.get('email_notifications'))
        user.updated_at = datetime.utcnow()
        
        try:
            db.session.commit()
            flash('Profile updated successfully!', 'success')
            return redirect(url_for('my_profile'))
        except Exception as e:
            db.session.rollback()
            flash('Error updating profile. Please try again.', 'error')
    
    return render_template('edit_profile.html', user=user)

@app.route('/delete-profile', methods=['POST'])
@login_required
def delete_profile():
    """Delete user profile (keeps forecast data but removes personal info)"""
    user = current_user
    
    try:
        # Instead of deleting the user, we'll anonymize their data
        # This preserves forecast data integrity while removing personal info
        user.username = f"deleted_user_{user.id}"
        user.email = None
        user.full_name = "Deleted User"
        user.theme_preference = "light"
        user.email_notifications = False
        
        db.session.commit()
        
        # Log out the user
        logout_user()
        
        flash('Your profile has been deleted. Your forecast data has been preserved for data integrity.', 'info')
        return redirect(url_for('index'))
        
    except Exception as e:
        db.session.rollback()
        flash('Error deleting profile. Please try again.', 'error')
        return redirect(url_for('edit_profile'))

@app.route('/user/<username>')
def view_user(username):
    """View another user's profile"""
    user = User.query.filter_by(username=username).first()
    if not user:
        flash('User not found.', 'error')
        return redirect(url_for('user_activity'))
    
    # Get user statistics
    user_forecasts = EventForecast.query.filter_by(created_by=username).all()
    total_forecasts = len([f for f in user_forecasts if f.revenue or f.adr or f.occupancy])
    
    # Get unique hotels this user has worked on
    unique_hotels = len(set([f.hotel_id for f in user_forecasts]))
    
    # Count uploads (estimate based on HotelActuals)
    total_uploads = HotelActuals.query.filter_by(created_by=username).count()
    
    # Get recent forecast activity (last 10)
    recent_forecasts = db.session.query(EventForecast, Event, Hotel).join(
        Event, EventForecast.event_id == Event.id
    ).join(
        Hotel, EventForecast.hotel_id == Hotel.id
    ).filter(
        EventForecast.created_by == username
    ).filter(
        or_(EventForecast.revenue.isnot(None), 
            EventForecast.adr.isnot(None), 
            EventForecast.occupancy.isnot(None))
    ).order_by(EventForecast.updated_at.desc()).limit(10).all()
    
    return render_template('view_user.html', 
                         user=user,
                         total_forecasts=total_forecasts,
                         total_uploads=total_uploads,
                         unique_hotels=unique_hotels,
                         recent_forecasts=recent_forecasts)

@app.route('/my-profile')
@login_required
def my_profile():
    """View user's personal profile with hotel cards and completion rates"""
    user_name = current_user.username
    user_id = current_user.id
    
    # Get hotels from multiple sources
    hotel_sources = set()
    
    # 1. Hotels from event forecasts (case-insensitive username matching)
    event_forecast_hotels = db.session.query(Hotel.id).join(EventForecast).filter(
        or_(
            EventForecast.created_by.ilike(f'%{user_name}%'),
            EventForecast.created_by == user_name
        )
    ).distinct().all()
    hotel_sources.update([h[0] for h in event_forecast_hotels])
    
    # 2. Hotels from monthly forecasts (case-insensitive username matching)
    monthly_forecast_hotels = db.session.query(Hotel.id).join(MonthlyForecast).filter(
        or_(
            MonthlyForecast.created_by.ilike(f'%{user_name}%'),
            MonthlyForecast.created_by == user_name
        )
    ).distinct().all()
    hotel_sources.update([h[0] for h in monthly_forecast_hotels])
    
    # 3. Hotels from assignments
    assigned_hotels = db.session.query(Hotel.id).join(HotelAssignment).filter(
        HotelAssignment.user_id == user_id
    ).distinct().all()
    hotel_sources.update([h[0] for h in assigned_hotels])
    
    # Get all unique hotels
    user_hotels = Hotel.query.filter(Hotel.id.in_(hotel_sources)).all() if hotel_sources else []
    
    # Calculate completion rates and metrics for each hotel
    hotel_cards = []
    total_event_forecasts = 0
    total_monthly_forecasts = 0
    total_hotels_worked = len(user_hotels)
    
    for hotel in user_hotels:
        # Get all event forecasts for this hotel by this user (case-insensitive)
        event_forecasts = EventForecast.query.filter(
            EventForecast.hotel_id == hotel.id,
            or_(
                EventForecast.created_by.ilike(f'%{user_name}%'),
                EventForecast.created_by == user_name
            )
        ).all()
        
        # Get all monthly forecasts for this hotel by this user (case-insensitive)
        monthly_forecasts = MonthlyForecast.query.filter(
            MonthlyForecast.hotel_id == hotel.id,
            or_(
                MonthlyForecast.created_by.ilike(f'%{user_name}%'),
                MonthlyForecast.created_by == user_name
            )
        ).all()
        
        # Check if hotel is assigned to user
        is_assigned = HotelAssignment.query.filter(
            HotelAssignment.hotel_id == hotel.id,
            HotelAssignment.user_id == user_id
        ).first() is not None
        
        # Count completed forecasts (those with actual data)
        completed_event_forecasts = len([f for f in event_forecasts if f.revenue or f.adr or f.occupancy])
        completed_monthly_forecasts = len([f for f in monthly_forecasts if f.revenue or f.adr or f.occupancy])
        
        total_event_forecasts += completed_event_forecasts
        total_monthly_forecasts += completed_monthly_forecasts
        
        # Get potential forecast days (events this hotel participates in)
        hotel_events = Event.query.filter_by(city=hotel.city).count()
        potential_forecast_days = hotel_events * 15  # Assuming 15-day forecast periods
        
        # Calculate completion rate (event forecasts + monthly forecasts vs potential)
        total_completed = completed_event_forecasts + completed_monthly_forecasts
        total_potential = potential_forecast_days + 365  # 365 days for monthly forecasts
        completion_rate = (total_completed / total_potential * 100) if total_potential > 0 else 0
        
        # Get recent activity (last 30 days)
        recent_event_activity = db.session.query(EventForecast).filter(
            EventForecast.hotel_id == hotel.id,
            or_(
                EventForecast.created_by.ilike(f'%{user_name}%'),
                EventForecast.created_by == user_name
            ),
            EventForecast.updated_at >= datetime.now() - timedelta(days=30)
        ).count()
        
        recent_monthly_activity = db.session.query(MonthlyForecast).filter(
            MonthlyForecast.hotel_id == hotel.id,
            or_(
                MonthlyForecast.created_by.ilike(f'%{user_name}%'),
                MonthlyForecast.created_by == user_name
            ),
            MonthlyForecast.updated_at >= datetime.now() - timedelta(days=30)
        ).count()
        
        recent_activity = recent_event_activity + recent_monthly_activity
        
        hotel_cards.append({
            'hotel': hotel,
            'event_forecasts': completed_event_forecasts,
            'monthly_forecasts': completed_monthly_forecasts,
            'total_forecasts': total_completed,
            'completion_rate': min(completion_rate, 100),  # Cap at 100%
            'recent_activity': recent_activity,
            'hotel_events': hotel_events,
            'is_assigned': is_assigned
        })
    
    # Sort by completion rate (highest first)
    hotel_cards.sort(key=lambda x: x['completion_rate'], reverse=True)
    
    # Get additional user statistics
    total_assignments = HotelAssignment.query.filter_by(user_id=user_id).count()
    
    # Get recent comprehensive activity (last 30 days)
    recent_activities = []
    thirty_days_ago = datetime.now() - timedelta(days=30)
    
    # Recent event forecasts
    recent_event_forecasts = db.session.query(EventForecast, Hotel, Event).join(
        Hotel, EventForecast.hotel_id == Hotel.id
    ).join(
        Event, EventForecast.event_id == Event.id
    ).filter(
        or_(
            EventForecast.created_by.ilike(f'%{user_name}%'),
            EventForecast.created_by == user_name
        ),
        EventForecast.updated_at >= thirty_days_ago
    ).order_by(EventForecast.updated_at.desc()).limit(10).all()
    
    for forecast, hotel, event in recent_event_forecasts:
        recent_activities.append({
            'type': 'Event Forecast',
            'description': f'Updated forecast for {hotel.hotel_code} - {event.event_name}',
            'timestamp': forecast.updated_at,
            'hotel': hotel,
            'icon': 'chart-line'
        })
    
    # Recent monthly forecasts
    recent_monthly_forecasts = db.session.query(MonthlyForecast, Hotel).join(
        Hotel, MonthlyForecast.hotel_id == Hotel.id
    ).filter(
        or_(
            MonthlyForecast.created_by.ilike(f'%{user_name}%'),
            MonthlyForecast.created_by == user_name
        ),
        MonthlyForecast.updated_at >= thirty_days_ago
    ).order_by(MonthlyForecast.updated_at.desc()).limit(10).all()
    
    for forecast, hotel in recent_monthly_forecasts:
        recent_activities.append({
            'type': 'Monthly Forecast',
            'description': f'Updated monthly forecast for {hotel.hotel_code} on {forecast.forecast_date.strftime("%Y-%m-%d")}',
            'timestamp': forecast.updated_at,
            'hotel': hotel,
            'icon': 'calendar'
        })
    
    # Recent assignments
    recent_assignments = db.session.query(HotelAssignment, Hotel).join(
        Hotel, HotelAssignment.hotel_id == Hotel.id
    ).filter(
        HotelAssignment.user_id == user_id,
        HotelAssignment.created_at >= thirty_days_ago
    ).order_by(HotelAssignment.created_at.desc()).limit(5).all()
    
    for assignment, hotel in recent_assignments:
        recent_activities.append({
            'type': 'Hotel Assignment',
            'description': f'Assigned to {hotel.hotel_code} as {assignment.role}',
            'timestamp': assignment.created_at,
            'hotel': hotel,
            'icon': 'user-check'
        })
    
    # Sort all activities by timestamp
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:15]  # Keep only the 15 most recent
    
    return render_template('my_profile.html', 
                         user_name=user_name,
                         hotel_cards=hotel_cards,
                         total_event_forecasts=total_event_forecasts,
                         total_monthly_forecasts=total_monthly_forecasts,
                         total_hotels_worked=total_hotels_worked,
                         total_assignments=total_assignments,
                         recent_activities=recent_activities)

# Hotel Actuals Data Upload/Download Routes

@app.route('/download-template')
@login_required
def download_template():
    """Generate and download CSV template for hotel actuals data entry"""
    
    try:
        # Create simple CSV template data with just a few sample rows
        template_data = [
            ['Date', 'Hotel_Code', 'Hotel_Name', 'Revenue_TY', 'Room_Nights_TY', 'ADR_TY', 'STLY_Revenue', 'STLY_Room_Nights', 'STLY_ADR'],
            ['2025-01-15', 'HTL001', 'Sample Hotel 1', '10000', '50', '200.00', '9000', '45', '200.00'],
            ['2025-01-16', 'HTL001', 'Sample Hotel 1', '12000', '55', '218.18', '11000', '50', '220.00'],
            ['2025-01-15', 'HTL002', 'Sample Hotel 2', '15000', '75', '200.00', '14000', '70', '200.00'],
        ]
        
        # Create CSV response
        output = StringIO()
        writer = csv.writer(output)
        writer.writerows(template_data)
        
        response = Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={"Content-disposition": f"attachment; filename=hotel_actuals_template_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
        
        return response
    
    except Exception as e:
        flash(f'Error generating template: {str(e)}', 'error')
        return redirect(url_for('actuals_dashboard'))

@app.route('/upload-hotels-excel', methods=['GET', 'POST'])
def upload_hotels_excel():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                df = pd.read_excel(file)
                
                # Expected columns: Hotel Code, Hotel Name, City, Inventory, Hotel Link, Address Link
                required_columns = ['Hotel Code', 'Hotel Name', 'City', 'Inventory']
                
                if not all(col in df.columns for col in required_columns):
                    flash(f'Excel file must contain columns: {", ".join(required_columns)}', 'error')
                    return redirect(request.url)
                
                success_count = 0
                error_count = 0
                
                for _, row in df.iterrows():
                    try:
                        # Check if hotel already exists
                        existing_hotel = Hotel.query.filter_by(hotel_code=row['Hotel Code']).first()
                        
                        if existing_hotel:
                            # Update existing hotel
                            existing_hotel.hotel_name = row['Hotel Name']
                            existing_hotel.city = row['City']
                            existing_hotel.inventory = int(row['Inventory'])
                            if 'Hotel Link' in df.columns and pd.notna(row['Hotel Link']):
                                existing_hotel.hotel_link = row['Hotel Link']
                            if 'Address Link' in df.columns and pd.notna(row['Address Link']):
                                existing_hotel.address_link = row['Address Link']
                        else:
                            # Create new hotel
                            new_hotel = Hotel(
                                hotel_code=row['Hotel Code'],
                                hotel_name=row['Hotel Name'],
                                city=row['City'],
                                inventory=int(row['Inventory']),
                                hotel_link=row.get('Hotel Link') if pd.notna(row.get('Hotel Link')) else None,
                                address_link=row.get('Address Link') if pd.notna(row.get('Address Link')) else None
                            )
                            db.session.add(new_hotel)
                        
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        print(f"Error processing row: {e}")
                
                db.session.commit()
                flash(f'Hotels upload completed: {success_count} successful, {error_count} errors', 'success')
                return redirect(url_for('manage_hotels'))
                
            except Exception as e:
                flash(f'Error processing Excel file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)
    
    return render_template('upload_hotels_excel.html')

@app.route('/upload-events-excel', methods=['GET', 'POST'])
def upload_events_excel():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            try:
                df = pd.read_excel(file)
                
                # Expected columns: Event Name, Start Date, End Date, City
                required_columns = ['Event Name', 'Start Date', 'End Date', 'City']
                
                if not all(col in df.columns for col in required_columns):
                    flash(f'Excel file must contain columns: {", ".join(required_columns)}', 'error')
                    return redirect(request.url)
                
                success_count = 0
                error_count = 0
                
                for _, row in df.iterrows():
                    try:
                        # Parse dates
                        start_date = pd.to_datetime(row['Start Date']).date()
                        end_date = pd.to_datetime(row['End Date']).date()
                        
                        # Check if event already exists
                        existing_event = Event.query.filter_by(
                            event_name=row['Event Name'],
                            start_date=start_date,
                            city=row['City']
                        ).first()
                        
                        if existing_event:
                            # Update existing event
                            existing_event.end_date = end_date
                        else:
                            # Create new event
                            new_event = Event(
                                event_name=row['Event Name'],
                                start_date=start_date,
                                end_date=end_date,
                                city=row['City']
                            )
                            db.session.add(new_event)
                        
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        print(f"Error processing row: {e}")
                
                db.session.commit()
                flash(f'Events upload completed: {success_count} successful, {error_count} errors', 'success')
                return redirect(url_for('manage_events'))
                
            except Exception as e:
                flash(f'Error processing Excel file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)
    
    return render_template('upload_events_excel.html')

# Hotel Detail and Edit Routes  
@app.route('/hotel/<int:hotel_id>')
@login_required
def hotel_detail(hotel_id):
    """View detailed hotel information with comprehensive data"""
    hotel = Hotel.query.get_or_404(hotel_id)
    
    # Get hotel comments (top-level only)
    comments = Comment.query.filter_by(
        hotel_id=hotel_id, 
        parent_id=None
    ).order_by(Comment.created_at.desc()).all()
    
    # Get recent forecasting activity for this hotel
    recent_forecasts = db.session.query(EventForecast, Event).join(
        Event, EventForecast.event_id == Event.id
    ).filter(
        EventForecast.hotel_id == hotel_id
    ).filter(
        or_(EventForecast.revenue.isnot(None), 
            EventForecast.adr.isnot(None), 
            EventForecast.occupancy.isnot(None))
    ).order_by(EventForecast.updated_at.desc()).limit(5).all()
    
    # Get events in the same city
    city_events = Event.query.filter_by(city=hotel.city).order_by(Event.start_date.desc()).limit(5).all()
    
    # Get statistics
    total_forecasts = EventForecast.query.filter_by(hotel_id=hotel_id).filter(
        or_(EventForecast.revenue.isnot(None), 
            EventForecast.adr.isnot(None), 
            EventForecast.occupancy.isnot(None))
    ).count()
    
    total_events = len(city_events)
    
    # Get comprehensive hotel analytics (last 90 days)
    hotel_analytics = generate_hotel_analytics(hotel_id)
    
    # Get hotel actuals for charts (if available)
    try:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=90)
        
        actuals = HotelActuals.query.filter(
            HotelActuals.hotel_id == hotel.id,
            HotelActuals.date.between(start_date, end_date)
        ).order_by(HotelActuals.date.desc()).all()
        
        # Create performance charts for this hotel
        charts = None
        if actuals:
            actual_revenue_data = []
            actual_adr_data = []
            actual_occupancy_data = []
            baseline_revenue_data = []
            baseline_adr_data = []
            baseline_occupancy_data = []
            impact_revenue_data = []
            
            for actual in actuals:
                # Current year data (use correct column names)
                if actual.ty_revenue:
                    actual_revenue_data.append({'x': actual.date, 'y': actual.ty_revenue, 'hotel': hotel.hotel_code})
                if actual.ty_adr:
                    actual_adr_data.append({'x': actual.date, 'y': actual.ty_adr, 'hotel': hotel.hotel_code})
                
                # Calculate occupancy from room nights
                if actual.ty_room_nights and hotel.inventory:
                    occupancy = (actual.ty_room_nights / hotel.inventory * 100)
                    actual_occupancy_data.append({'x': actual.date, 'y': occupancy, 'hotel': hotel.hotel_code})
                
                # Baseline data (STLY)
                if actual.stly_revenue:
                    baseline_revenue_data.append({'x': actual.date, 'y': actual.stly_revenue, 'hotel': hotel.hotel_code})
                if actual.stly_adr:
                    baseline_adr_data.append({'x': actual.date, 'y': actual.stly_adr, 'hotel': hotel.hotel_code})
                
                # Calculate baseline occupancy from room nights
                if actual.stly_room_nights and hotel.inventory:
                    baseline_occupancy = (actual.stly_room_nights / hotel.inventory * 100)
                    baseline_occupancy_data.append({'x': actual.date, 'y': baseline_occupancy, 'hotel': hotel.hotel_code})
                
                # Calculate revenue impact
                if actual.ty_revenue and actual.stly_revenue:
                    revenue_impact = actual.ty_revenue - actual.stly_revenue
                    impact_revenue_data.append({'x': actual.date, 'y': revenue_impact, 'hotel': hotel.hotel_code})
            
            if actual_revenue_data or baseline_revenue_data:
                charts = create_performance_charts(
                    actual_revenue_data, baseline_revenue_data, impact_revenue_data,
                    actual_adr_data, baseline_adr_data,
                    actual_occupancy_data, baseline_occupancy_data
                )
    except Exception as e:
        print(f"Error generating charts: {e}")
        actuals = []
        charts = None
    
    return render_template('hotel_detail.html',
                         hotel=hotel,
                         comments=comments,
                         recent_forecasts=recent_forecasts,
                         city_events=city_events,
                         total_forecasts=total_forecasts,
                         total_events=total_events,
                         actuals=actuals,
                         charts=charts,
                         hotel_analytics=hotel_analytics)

# Event Detail Routes
@app.route('/event/<int:event_id>')
@login_required  
def event_detail(event_id):
    """View detailed event information with comprehensive data"""
    event = Event.query.get_or_404(event_id)
    
    # Get event comments (top-level only)
    comments = Comment.query.filter_by(
        event_id=event_id,
        parent_id=None
    ).order_by(Comment.created_at.desc()).all()
    
    # Get hotels in the same city
    city_hotels = Hotel.query.filter_by(city=event.city).all()
    
    # Get recent forecasts for this event
    recent_forecasts = db.session.query(EventForecast, Hotel).join(
        Hotel, EventForecast.hotel_id == Hotel.id
    ).filter(
        EventForecast.event_id == event_id
    ).order_by(EventForecast.updated_at.desc()).limit(10).all()
    
    # Get other events in the same city
    other_events = Event.query.filter(
        Event.city == event.city,
        Event.id != event_id
    ).order_by(Event.start_date.desc()).limit(5).all()
    
    return render_template('event_detail.html',
                         event=event,
                         comments=comments, 
                         recent_forecasts=recent_forecasts,
                         city_hotels=city_hotels,
                         other_events=other_events)

def create_performance_charts(actual_revenue, baseline_revenue, impact_revenue, 
                            actual_adr, baseline_adr, actual_occupancy, baseline_occupancy):
    """Create Chart.js compatible data for performance charts"""
    charts = {}
    
    # Revenue Comparison Chart
    if actual_revenue or baseline_revenue:
        revenue_data = {
            'type': 'line',
            'data': {
                'datasets': []
            },
            'options': {
                'responsive': True,
                'maintainAspectRatio': False,
                'plugins': {
                    'title': {
                        'display': True,
                        'text': 'Revenue Performance: Actual vs Baseline'
                    }
                },
                'scales': {
                    'y': {
                        'beginAtZero': False,
                        'title': {
                            'display': True,
                            'text': 'Revenue ($)'
                        }
                    },
                    'x': {
                        'title': {
                            'display': True,
                            'text': 'Date'
                        }
                    }
                }
            }
        }
        
        if actual_revenue:
            revenue_data['data']['datasets'].append({
                'label': 'Actual Revenue',
                'data': [{'x': item['x'].strftime('%Y-%m-%d'), 'y': item['y']} for item in actual_revenue],
                'borderColor': '#28a745',
                'backgroundColor': '#28a74520',
                'tension': 0.4
            })
        
        if baseline_revenue:
            revenue_data['data']['datasets'].append({
                'label': 'Baseline (STLY)',
                'data': [{'x': item['x'].strftime('%Y-%m-%d'), 'y': item['y']} for item in baseline_revenue],
                'borderColor': '#6c757d',
                'backgroundColor': '#6c757d20',
                'borderDash': [5, 5],
                'tension': 0.4
            })
        
        charts['revenue_chart'] = revenue_data
    
    # ADR Comparison Chart
    if actual_adr or baseline_adr:
        adr_data = {
            'type': 'line',
            'data': {
                'datasets': []
            },
            'options': {
                'responsive': True,
                'maintainAspectRatio': False,
                'plugins': {
                    'title': {
                        'display': True,
                        'text': 'ADR Performance: Actual vs Baseline'
                    }
                },
                'scales': {
                    'y': {
                        'beginAtZero': False,
                        'title': {
                            'display': True,
                            'text': 'Average Daily Rate ($)'
                        }
                    },
                    'x': {
                        'title': {
                            'display': True,
                            'text': 'Date'
                        }
                    }
                }
            }
        }
        
        if actual_adr:
            adr_data['data']['datasets'].append({
                'label': 'Actual ADR',
                'data': [{'x': item['x'].strftime('%Y-%m-%d'), 'y': item['y']} for item in actual_adr],
                'borderColor': '#007bff',
                'backgroundColor': '#007bff20',
                'tension': 0.4
            })
        
        if baseline_adr:
            adr_data['data']['datasets'].append({
                'label': 'Baseline ADR (STLY)',
                'data': [{'x': item['x'].strftime('%Y-%m-%d'), 'y': item['y']} for item in baseline_adr],
                'borderColor': '#6c757d',
                'backgroundColor': '#6c757d20',
                'borderDash': [5, 5],
                'tension': 0.4
            })
        
        charts['adr_chart'] = adr_data
    
    # Occupancy Comparison Chart
    if actual_occupancy or baseline_occupancy:
        occupancy_data = {
            'type': 'line',
            'data': {
                'datasets': []
            },
            'options': {
                'responsive': True,
                'maintainAspectRatio': False,
                'plugins': {
                    'title': {
                        'display': True,
                        'text': 'Occupancy Performance: Actual vs Baseline'
                    }
                },
                'scales': {
                    'y': {
                        'beginAtZero': True,
                        'max': 100,
                        'title': {
                            'display': True,
                            'text': 'Occupancy (%)'
                        }
                    },
                    'x': {
                        'title': {
                            'display': True,
                            'text': 'Date'
                        }
                    }
                }
            }
        }
        
        if actual_occupancy:
            occupancy_data['data']['datasets'].append({
                'label': 'Actual Occupancy',
                'data': [{'x': item['x'].strftime('%Y-%m-%d'), 'y': item['y']} for item in actual_occupancy],
                'borderColor': '#17a2b8',
                'backgroundColor': '#17a2b820',
                'tension': 0.4
            })
        
        if baseline_occupancy:
            occupancy_data['data']['datasets'].append({
                'label': 'Baseline Occupancy (STLY)',
                'data': [{'x': item['x'].strftime('%Y-%m-%d'), 'y': item['y']} for item in baseline_occupancy],
                'borderColor': '#6c757d',
                'backgroundColor': '#6c757d20',
                'borderDash': [5, 5],
                'tension': 0.4
            })
        
        charts['occupancy_chart'] = occupancy_data
    
    # Impact Analysis Chart
    if impact_revenue:
        impact_data = {
            'type': 'bar',
            'data': {
                'labels': [item['x'].strftime('%Y-%m-%d') for item in impact_revenue],
                'datasets': [{
                    'label': 'Revenue Impact vs STLY',
                    'data': [item['y'] for item in impact_revenue],
                    'backgroundColor': [
                        '#28a745' if val > 0 else '#dc3545' if val < 0 else '#6c757d'
                        for val in [item['y'] for item in impact_revenue]
                    ],
                    'borderWidth': 1
                }]
            },
            'options': {
                'responsive': True,
                'maintainAspectRatio': False,
                'plugins': {
                    'title': {
                        'display': True,
                        'text': 'Revenue Impact Analysis (vs STLY)'
                    },
                    'legend': {
                        'display': False
                    }
                },
                'scales': {
                    'y': {
                        'title': {
                            'display': True,
                            'text': 'Revenue Impact ($)'
                        }
                    },
                    'x': {
                        'title': {
                            'display': True,
                            'text': 'Date'
                        }
                    }
                }
            }
        }
        
        charts['impact_chart'] = impact_data
    
    return charts

# Task Management Routes
@app.route('/tasks')
@login_required
def tasks():
    """Main task management page"""
    # Get tasks assigned to current user
    my_tasks = Task.query.filter_by(assigned_to_id=current_user.id).order_by(
        Task.due_date.asc().nullslast(), Task.created_at.desc()
    ).all()
    
    # Get tasks assigned by current user
    assigned_tasks = Task.query.filter_by(assigned_by_id=current_user.id).order_by(
        Task.created_at.desc()
    ).all()
    
    # Get tasks from users being followed
    followed_user_ids = [follow.followed_user_id for follow in current_user.following_users]
    following_tasks = []
    if followed_user_ids:
        following_tasks = Task.query.filter(
            or_(
                and_(Task.assigned_to_id.in_(followed_user_ids), 
                     UserTaskFollow.follow_assigned_tasks == True),
                and_(Task.assigned_by_id.in_(followed_user_ids), 
                     UserTaskFollow.follow_created_tasks == True)
            )
        ).order_by(Task.created_at.desc()).all()
    
    # Get all tasks (for admin/manager view)
    all_tasks = Task.query.order_by(Task.created_at.desc()).all()
    
    return render_template('tasks.html',
                         my_tasks=my_tasks,
                         assigned_tasks=assigned_tasks,
                         following_tasks=following_tasks,
                         all_tasks=all_tasks)

@app.route('/tasks/create', methods=['GET', 'POST'])
@login_required
def create_task():
    """Create a new task"""
    if request.method == 'POST':
        try:
            # Parse due date if provided
            due_date = None
            if request.form.get('due_date') and request.form.get('due_time'):
                due_date_str = f"{request.form.get('due_date')} {request.form.get('due_time')}"
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d %H:%M')
            
            # Create new task
            task = Task(
                title=request.form.get('title'),
                description=request.form.get('description'),
                priority=request.form.get('priority', 'medium'),
                category=request.form.get('category', 'general'),
                assigned_by_id=current_user.id,
                assigned_to_id=int(request.form.get('assigned_to_id')),
                hotel_id=int(request.form.get('hotel_id')) if request.form.get('hotel_id') else None,
                event_id=int(request.form.get('event_id')) if request.form.get('event_id') else None,
                due_date=due_date
            )
            
            db.session.add(task)
            db.session.commit()
            
            flash(f'Task "{task.title}" created successfully!', 'success')
            return redirect(url_for('task_detail', task_id=task.id))
            
        except Exception as e:
            db.session.rollback()
            flash('Error creating task. Please try again.', 'error')
    
    # Get users for assignment dropdown
    users = User.query.filter_by(is_active=True).order_by(User.full_name, User.username).all()
    hotels = Hotel.query.order_by(Hotel.hotel_name).all()
    events = Event.query.filter(Event.end_date >= datetime.now().date()).order_by(Event.start_date).all()
    
    return render_template('create_task.html', users=users, hotels=hotels, events=events)

@app.route('/tasks/<int:task_id>')
@login_required
def task_detail(task_id):
    """View task details"""
    task = Task.query.get_or_404(task_id)
    
    # Get task comments
    comments = TaskComment.query.filter_by(task_id=task_id).order_by(TaskComment.created_at.desc()).all()
    
    return render_template('task_detail.html', task=task, comments=comments)

@app.route('/tasks/<int:task_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_task(task_id):
    """Edit a task"""
    task = Task.query.get_or_404(task_id)
    
    # Check permissions - only creator or assignee can edit
    if current_user.id not in [task.assigned_by_id, task.assigned_to_id]:
        flash('You do not have permission to edit this task.', 'error')
        return redirect(url_for('task_detail', task_id=task_id))
    
    if request.method == 'POST':
        try:
            # Parse due date if provided
            due_date = None
            if request.form.get('due_date') and request.form.get('due_time'):
                due_date_str = f"{request.form.get('due_date')} {request.form.get('due_time')}"
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d %H:%M')
            
            # Update task
            task.title = request.form.get('title')
            task.description = request.form.get('description')
            task.priority = request.form.get('priority', 'medium')
            task.category = request.form.get('category', 'general')
            task.assigned_to_id = int(request.form.get('assigned_to_id'))
            task.hotel_id = int(request.form.get('hotel_id')) if request.form.get('hotel_id') else None
            task.event_id = int(request.form.get('event_id')) if request.form.get('event_id') else None
            task.due_date = due_date
            
            db.session.commit()
            
            flash(f'Task "{task.title}" updated successfully!', 'success')
            return redirect(url_for('task_detail', task_id=task_id))
            
        except Exception as e:
            db.session.rollback()
            flash('Error updating task. Please try again.', 'error')
    
    # Get users for assignment dropdown
    users = User.query.filter_by(is_active=True).order_by(User.full_name, User.username).all()
    hotels = Hotel.query.order_by(Hotel.hotel_name).all()
    events = Event.query.filter(Event.end_date >= datetime.now().date()).order_by(Event.start_date).all()
    
    return render_template('edit_task.html', task=task, users=users, hotels=hotels, events=events)

@app.route('/tasks/<int:task_id>/status', methods=['POST'])
@login_required
def update_task_status(task_id):
    """Update task status"""
    task = Task.query.get_or_404(task_id)
    new_status = request.form.get('status')
    
    # Check permissions - only assignee can update status
    if current_user.id != task.assigned_to_id:
        flash('You can only update the status of tasks assigned to you.', 'error')
        return redirect(url_for('task_detail', task_id=task_id))
    
    try:
        old_status = task.status
        task.status = new_status
        
        # Set completion timestamp if marking as completed
        if new_status == 'completed':
            task.completed_at = datetime.utcnow()
        
        # Add status change comment
        comment = TaskComment(
            task_id=task_id,
            user_id=current_user.id,
            comment=f"Status changed from {old_status.replace('_', ' ').title()} to {new_status.replace('_', ' ').title()}",
            old_status=old_status,
            new_status=new_status
        )
        
        db.session.add(comment)
        db.session.commit()
        
        flash(f'Task status updated to {new_status.replace("_", " ").title()}!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('Error updating task status. Please try again.', 'error')
    
    return redirect(url_for('task_detail', task_id=task_id))

@app.route('/tasks/<int:task_id>/comment', methods=['POST'])
@login_required
def add_task_comment(task_id):
    """Add a comment to a task"""
    task = Task.query.get_or_404(task_id)
    comment_text = request.form.get('comment')
    
    if comment_text:
        try:
            comment = TaskComment(
                task_id=task_id,
                user_id=current_user.id,
                comment=comment_text
            )
            
            db.session.add(comment)
            db.session.commit()
            
            flash('Comment added successfully!', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash('Error adding comment. Please try again.', 'error')
    
    return redirect(url_for('task_detail', task_id=task_id))

@app.route('/calendar')
@login_required
def calendar():
    """Task calendar view"""
    # Get tasks for current user
    my_tasks = Task.query.filter_by(assigned_to_id=current_user.id).filter(
        Task.due_date.isnot(None)
    ).all()
    
    # Get tasks assigned by current user
    assigned_tasks = Task.query.filter_by(assigned_by_id=current_user.id).filter(
        Task.due_date.isnot(None)
    ).all()
    
    # Get followed tasks
    followed_user_ids = [follow.followed_user_id for follow in current_user.following_users]
    following_tasks = []
    if followed_user_ids:
        following_tasks = Task.query.filter(
            or_(Task.assigned_to_id.in_(followed_user_ids), Task.assigned_by_id.in_(followed_user_ids))
        ).filter(Task.due_date.isnot(None)).all()
    
    # Convert tasks to calendar events
    calendar_events = []
    
    # My tasks (blue)
    for task in my_tasks:
        calendar_events.append({
            'id': f'my-{task.id}',
            'title': task.title,
            'start': task.due_date.isoformat(),
            'backgroundColor': '#007bff',
            'borderColor': '#007bff',
            'url': url_for('task_detail', task_id=task.id),
            'extendedProps': {
                'type': 'assigned_to_me',
                'priority': task.priority,
                'status': task.status,
                'assignedBy': task.assigned_by.get_display_name()
            }
        })
    
    # Tasks I assigned (green)
    for task in assigned_tasks:
        if task not in my_tasks:  # Avoid duplicates for self-assigned tasks
            calendar_events.append({
                'id': f'assigned-{task.id}',
                'title': f'{task.title} (assigned to {task.assigned_to.get_display_name()})',
                'start': task.due_date.isoformat(),
                'backgroundColor': '#28a745',
                'borderColor': '#28a745',
                'url': url_for('task_detail', task_id=task.id),
                'extendedProps': {
                    'type': 'assigned_by_me',
                    'priority': task.priority,
                    'status': task.status,
                    'assignedTo': task.assigned_to.get_display_name()
                }
            })
    
    # Following tasks (orange)
    for task in following_tasks:
        if task not in my_tasks and task not in assigned_tasks:  # Avoid duplicates
            calendar_events.append({
                'id': f'following-{task.id}',
                'title': f'{task.title} ({task.assigned_to.get_display_name()})',
                'start': task.due_date.isoformat(),
                'backgroundColor': '#fd7e14',
                'borderColor': '#fd7e14',
                'url': url_for('task_detail', task_id=task.id),
                'extendedProps': {
                    'type': 'following',
                    'priority': task.priority,
                    'status': task.status,
                    'assignedBy': task.assigned_by.get_display_name(),
                    'assignedTo': task.assigned_to.get_display_name()
                }
            })
    
    return render_template('calendar.html', calendar_events=calendar_events)

@app.route('/follows')
@login_required
def manage_follows():
    """Manage user follows for task tracking"""
    # Get all active users except current user
    all_users = User.query.filter(
        User.is_active == True,
        User.id != current_user.id
    ).order_by(User.full_name, User.username).all()
    
    # Get current follows
    current_follows = {follow.followed_user_id: follow for follow in current_user.following_users}
    
    return render_template('manage_follows.html', all_users=all_users, current_follows=current_follows)

@app.route('/follows/toggle', methods=['POST'])
@login_required
def toggle_follow():
    """Toggle following a user"""
    user_id = int(request.form.get('user_id'))
    follow_assigned = request.form.get('follow_assigned') == 'on'
    follow_created = request.form.get('follow_created') == 'on'
    
    # Check if already following
    existing_follow = UserTaskFollow.query.filter_by(
        follower_id=current_user.id,
        followed_user_id=user_id
    ).first()
    
    try:
        if existing_follow:
            if follow_assigned or follow_created:
                # Update follow preferences
                existing_follow.follow_assigned_tasks = follow_assigned
                existing_follow.follow_created_tasks = follow_created
                action = 'updated'
            else:
                # Remove follow if neither option selected
                db.session.delete(existing_follow)
                action = 'removed'
        else:
            if follow_assigned or follow_created:
                # Create new follow
                new_follow = UserTaskFollow(
                    follower_id=current_user.id,
                    followed_user_id=user_id,
                    follow_assigned_tasks=follow_assigned,
                    follow_created_tasks=follow_created
                )
                db.session.add(new_follow)
                action = 'added'
            else:
                action = 'no_change'
        
        db.session.commit()
        
        if action != 'no_change':
            user = User.query.get(user_id)
            flash(f'Follow settings {action} for {user.get_display_name()}!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash('Error updating follow settings. Please try again.', 'error')
    
    return redirect(url_for('manage_follows'))

# Hotel Assignment Routes
@app.route('/hotels/<int:hotel_id>/assign', methods=['POST'])
@login_required
def assign_hotel(hotel_id):
    """Assign a hotel to a user"""
    hotel = Hotel.query.get_or_404(hotel_id)
    user_id = request.form.get('user_id')
    role = request.form.get('role', 'manager')
    is_primary = request.form.get('is_primary') == 'on'
    
    if not user_id:
        flash('Please select a user to assign the hotel to.', 'error')
        return redirect(request.referrer or url_for('dashboard'))
    
    try:
        # Check if assignment already exists
        existing_assignment = HotelAssignment.query.filter_by(
            hotel_id=hotel_id,
            user_id=int(user_id)
        ).first()
        
        if existing_assignment:
            if existing_assignment.is_active:
                flash(f'Hotel "{hotel.hotel_name}" is already assigned to this user.', 'warning')
            else:
                # Reactivate existing assignment
                existing_assignment.is_active = True
                existing_assignment.role = role
                existing_assignment.is_primary = is_primary
                existing_assignment.assigned_by_id = current_user.id
                existing_assignment.updated_at = datetime.utcnow()
                db.session.commit()
                
                user = User.query.get(int(user_id))
                flash(f'Hotel "{hotel.hotel_name}" successfully reassigned to {user.get_display_name()}!', 'success')
        else:
            # Create new assignment
            assignment = HotelAssignment(
                hotel_id=hotel_id,
                user_id=int(user_id),
                assigned_by_id=current_user.id,
                role=role,
                is_primary=is_primary
            )
            
            db.session.add(assignment)
            db.session.commit()
            
            user = User.query.get(int(user_id))
            flash(f'Hotel "{hotel.hotel_name}" successfully assigned to {user.get_display_name()}!', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash('Error assigning hotel. Please try again.', 'error')
    
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/hotels/<int:hotel_id>/unassign/<int:user_id>', methods=['POST'])
@login_required
def unassign_hotel(hotel_id, user_id):
    """Remove hotel assignment from a user"""
    assignment = HotelAssignment.query.filter_by(
        hotel_id=hotel_id,
        user_id=user_id,
        is_active=True
    ).first_or_404()
    
    try:
        assignment.is_active = False
        assignment.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash(f'Hotel assignment removed successfully!', 'success')
    
    except Exception as e:
        db.session.rollback()
        flash('Error removing hotel assignment. Please try again.', 'error')
    
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/assignments')
@login_required
def manage_assignments():
    """View and manage all hotel assignments"""
    # Get all active assignments
    assignments = db.session.query(HotelAssignment, Hotel, User).join(
        Hotel, HotelAssignment.hotel_id == Hotel.id
    ).join(
        User, HotelAssignment.user_id == User.id
    ).filter(
        HotelAssignment.is_active == True
    ).order_by(Hotel.hotel_name, User.full_name).all()
    
    # Get all hotels and users for new assignments
    hotels = Hotel.query.order_by(Hotel.hotel_name).all()
    users = User.query.filter_by(is_active=True).order_by(User.full_name, User.username).all()
    
    return render_template('assignments.html', 
                         assignments=assignments,
                         hotels=hotels,
                         users=users)

# Advanced Analytics and Reporting Routes
@app.route('/analytics')
@login_required
def analytics_dashboard():
    """Advanced analytics and reporting dashboard"""
    # Get date range for analysis (default to last 90 days)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    hotel_filter = request.args.get('hotel_filter')  # Optional hotel filter
    
    if not date_from:
        date_from = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = datetime.now().strftime('%Y-%m-%d')
    
    # Convert to datetime objects
    start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
    end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Get all hotels for filter dropdown
    hotels = Hotel.query.order_by(Hotel.hotel_code).all()
    
    # Get comprehensive analytics data
    analytics_data = generate_analytics_data(start_date, end_date, hotel_filter)
    
    return render_template('analytics_dashboard.html',
                         analytics_data=analytics_data,
                         hotels=hotels,
                         selected_hotel=hotel_filter,
                         date_from=date_from,
                         date_to=date_to)

def generate_analytics_data(start_date, end_date, hotel_filter=None):
    """Generate comprehensive analytics data for the dashboard"""
    analytics = {}
    
    # 1. Portfolio Performance Overview
    portfolio_stats = calculate_portfolio_performance(start_date, end_date, hotel_filter)
    analytics['portfolio'] = portfolio_stats
    
    # 2. Hotel Performance Rankings
    hotel_rankings = calculate_hotel_rankings(start_date, end_date, hotel_filter)
    analytics['hotel_rankings'] = hotel_rankings
    
    # 3. Forecast Accuracy Analysis
    forecast_accuracy = calculate_forecast_accuracy(start_date, end_date, hotel_filter)
    analytics['forecast_accuracy'] = forecast_accuracy
    
    # 4. Revenue Trend Analysis
    revenue_trends = calculate_revenue_trends(start_date, end_date, hotel_filter)
    analytics['revenue_trends'] = revenue_trends
    
    # 5. Occupancy and ADR Analysis
    occupancy_adr_analysis = calculate_occupancy_adr_trends(start_date, end_date, hotel_filter)
    analytics['occupancy_adr'] = occupancy_adr_analysis
    
    # 6. Event Impact Analysis
    event_impact = calculate_event_impact_analysis(start_date, end_date, hotel_filter)
    analytics['event_impact'] = event_impact
    
    # 7. User Activity and Productivity
    user_productivity = calculate_user_productivity(start_date, end_date)
    analytics['user_productivity'] = user_productivity
    
    # 8. Market Comparison and Benchmarking
    market_comparison = calculate_market_comparison(start_date, end_date)
    analytics['market_comparison'] = market_comparison
    
    return analytics

def calculate_portfolio_performance(start_date, end_date, hotel_filter=None):
    """Calculate overall portfolio performance metrics"""
    # Get actual performance data
    actuals = db.session.query(HotelActuals, Hotel).join(
        Hotel, HotelActuals.hotel_id == Hotel.id
    ).filter(
        HotelActuals.date.between(start_date, end_date)
    ).all()
    
    total_revenue_ty = sum(actual.ty_revenue for actual, hotel in actuals if actual.ty_revenue)
    total_revenue_stly = sum(actual.stly_revenue for actual, hotel in actuals if actual.stly_revenue)
    total_room_nights_ty = sum(actual.ty_room_nights for actual, hotel in actuals if actual.ty_room_nights)
    total_room_nights_stly = sum(actual.stly_room_nights for actual, hotel in actuals if actual.stly_room_nights)
    
    # Calculate key metrics
    revenue_growth = ((total_revenue_ty - total_revenue_stly) / total_revenue_stly * 100) if total_revenue_stly else 0
    room_nights_growth = ((total_room_nights_ty - total_room_nights_stly) / total_room_nights_stly * 100) if total_room_nights_stly else 0
    
    avg_adr_ty = total_revenue_ty / total_room_nights_ty if total_room_nights_ty else 0
    avg_adr_stly = total_revenue_stly / total_room_nights_stly if total_room_nights_stly else 0
    adr_growth = ((avg_adr_ty - avg_adr_stly) / avg_adr_stly * 100) if avg_adr_stly else 0
    
    # Calculate total inventory and occupancy
    total_inventory = sum(hotel.inventory for actual, hotel in actuals)
    days_in_period = (end_date - start_date).days + 1
    total_available_rooms = total_inventory * days_in_period if total_inventory else 1
    
    occupancy_ty = (total_room_nights_ty / total_available_rooms * 100) if total_available_rooms else 0
    occupancy_stly = (total_room_nights_stly / total_available_rooms * 100) if total_available_rooms else 0
    occupancy_growth = occupancy_ty - occupancy_stly
    
    return {
        'total_revenue_ty': total_revenue_ty,
        'total_revenue_stly': total_revenue_stly,
        'revenue_growth': revenue_growth,
        'total_room_nights_ty': total_room_nights_ty,
        'total_room_nights_stly': total_room_nights_stly,
        'room_nights_growth': room_nights_growth,
        'avg_adr_ty': avg_adr_ty,
        'avg_adr_stly': avg_adr_stly,
        'adr_growth': adr_growth,
        'occupancy_ty': occupancy_ty,
        'occupancy_stly': occupancy_stly,
        'occupancy_growth': occupancy_growth,
        'total_hotels': len(set(hotel.id for actual, hotel in actuals))
    }

def calculate_hotel_rankings(start_date, end_date, hotel_filter=None):
    """Calculate hotel performance rankings"""
    # Get hotel performance data
    hotel_performance = db.session.query(
        Hotel,
        db.func.sum(HotelActuals.ty_revenue).label('total_revenue'),
        db.func.sum(HotelActuals.stly_revenue).label('baseline_revenue'),
        db.func.sum(HotelActuals.ty_room_nights).label('total_room_nights'),
        db.func.avg(HotelActuals.ty_adr).label('avg_adr'),
        db.func.count(HotelActuals.id).label('data_points')
    ).join(
        HotelActuals, Hotel.id == HotelActuals.hotel_id
    ).filter(
        HotelActuals.date.between(start_date, end_date)
    ).group_by(Hotel.id).all()
    
    rankings = []
    for hotel, total_rev, baseline_rev, room_nights, avg_adr, data_points in hotel_performance:
        revenue_growth = ((total_rev - baseline_rev) / baseline_rev * 100) if baseline_rev else 0
        occupancy = (room_nights / (hotel.inventory * data_points) * 100) if hotel.inventory and data_points else 0
        
        rankings.append({
            'hotel': hotel,
            'total_revenue': total_rev or 0,
            'revenue_growth': revenue_growth,
            'occupancy': occupancy,
            'avg_adr': avg_adr or 0,
            'room_nights': room_nights or 0,
            'data_points': data_points
        })
    
    # Sort by revenue growth
    rankings.sort(key=lambda x: x['revenue_growth'], reverse=True)
    
    return rankings[:10]  # Top 10 performers

def calculate_forecast_accuracy(start_date, end_date, hotel_filter=None):
    """Calculate forecasting accuracy metrics"""
    # Simplified approach - get all forecasts and match with actuals
    forecasts = db.session.query(EventForecast, Hotel, Event).join(
        Hotel, EventForecast.hotel_id == Hotel.id
    ).join(
        Event, EventForecast.event_id == Event.id
    ).filter(
        EventForecast.revenue.isnot(None)
    ).all()
    
    accuracy_data = []
    revenue_errors = []
    adr_errors = []
    occupancy_errors = []
    
    for forecast, hotel, event in forecasts:
        # Use event dates since EventForecast doesn't have day_number
        # Check if event overlaps with our date range
        if event.end_date < start_date or event.start_date > end_date:
            continue
        
        # Use event start date as forecast reference
        forecast_date = event.start_date
            
        # Find corresponding actual data
        actual = HotelActuals.query.filter(
            HotelActuals.hotel_id == hotel.id,
            HotelActuals.date == forecast_date,
            HotelActuals.ty_revenue.isnot(None)
        ).first()
        
        if not actual:
            continue
        
        # Revenue accuracy
        if forecast.revenue and actual.ty_revenue:
            revenue_error = abs(forecast.revenue - actual.ty_revenue) / actual.ty_revenue * 100
            revenue_errors.append(revenue_error)
        
        # ADR accuracy
        if forecast.adr and actual.ty_adr:
            adr_error = abs(forecast.adr - actual.ty_adr) / actual.ty_adr * 100
            adr_errors.append(adr_error)
        
        # Occupancy accuracy
        if forecast.occupancy and actual.ty_room_nights and hotel.inventory:
            actual_occupancy = actual.ty_room_nights / hotel.inventory * 100
            occupancy_error = abs(forecast.occupancy - actual_occupancy) / actual_occupancy * 100
            occupancy_errors.append(occupancy_error)
        
        accuracy_data.append({
            'hotel': hotel,
            'event': event,
            'forecast_date': forecast_date,
            'forecast_revenue': forecast.revenue,
            'actual_revenue': actual.ty_revenue,
            'forecast_adr': forecast.adr,
            'actual_adr': actual.ty_adr,
            'forecast_occupancy': forecast.occupancy,
            'actual_occupancy': actual.ty_room_nights / hotel.inventory * 100 if hotel.inventory else 0
        })
    
    return {
        'total_forecasts': len(accuracy_data),
        'avg_revenue_accuracy': 100 - (sum(revenue_errors) / len(revenue_errors)) if revenue_errors else 0,
        'avg_adr_accuracy': 100 - (sum(adr_errors) / len(adr_errors)) if adr_errors else 0,
        'avg_occupancy_accuracy': 100 - (sum(occupancy_errors) / len(occupancy_errors)) if occupancy_errors else 0,
        'detailed_data': accuracy_data[:20]  # Top 20 recent comparisons
    }

def calculate_revenue_trends(start_date, end_date, hotel_filter=None):
    """Calculate revenue trend analysis"""
    # Build base query
    query = db.session.query(
        HotelActuals.date,
        db.func.sum(HotelActuals.ty_revenue).label('total_revenue'),
        db.func.sum(HotelActuals.stly_revenue).label('baseline_revenue'),
        db.func.count(HotelActuals.id).label('hotel_count')
    ).filter(
        HotelActuals.date.between(start_date, end_date),
        HotelActuals.ty_revenue.isnot(None)
    )
    
    # Add hotel filter if specified
    if hotel_filter:
        hotel = Hotel.query.filter_by(hotel_code=hotel_filter).first()
        if hotel:
            query = query.filter(HotelActuals.hotel_id == hotel.id)
    
    daily_revenue = query.group_by(HotelActuals.date).order_by(HotelActuals.date).all()
    
    # Calculate weekly moving averages
    weekly_data = []
    for i in range(len(daily_revenue) - 6):
        week_data = daily_revenue[i:i+7]
        avg_revenue = sum(day.total_revenue or 0 for day in week_data) / 7
        avg_baseline = sum(day.baseline_revenue or 0 for day in week_data) / 7
        
        weekly_data.append({
            'date': week_data[3].date,  # Middle date of the week
            'avg_revenue': avg_revenue,
            'avg_baseline': avg_baseline,
            'growth_rate': ((avg_revenue - avg_baseline) / avg_baseline * 100) if avg_baseline else 0
        })
    
    return {
        'daily_data': [{'date': day.date, 'revenue': day.total_revenue, 'baseline': day.baseline_revenue} for day in daily_revenue],
        'weekly_trends': weekly_data,
        'total_days': len(daily_revenue)
    }

def calculate_occupancy_adr_trends(start_date, end_date, hotel_filter=None):
    """Calculate occupancy and ADR trend analysis"""
    # Build base query
    query = db.session.query(
        HotelActuals.date,
        db.func.sum(HotelActuals.ty_room_nights).label('total_room_nights'),
        db.func.sum(HotelActuals.stly_room_nights).label('baseline_room_nights'),
        db.func.avg(HotelActuals.ty_adr).label('avg_adr'),
        db.func.avg(HotelActuals.stly_adr).label('baseline_adr'),
        db.func.sum(Hotel.inventory).label('total_inventory')
    ).join(
        Hotel, HotelActuals.hotel_id == Hotel.id
    ).filter(
        HotelActuals.date.between(start_date, end_date),
        HotelActuals.ty_room_nights.isnot(None)
    )
    
    # Add hotel filter if specified
    if hotel_filter:
        hotel = Hotel.query.filter_by(hotel_code=hotel_filter).first()
        if hotel:
            query = query.filter(HotelActuals.hotel_id == hotel.id)
    
    daily_data = query.group_by(HotelActuals.date).order_by(HotelActuals.date).all()
    
    occupancy_trends = []
    adr_trends = []
    
    for day in daily_data:
        occupancy = (day.total_room_nights / day.total_inventory * 100) if day.total_inventory else 0
        baseline_occupancy = (day.baseline_room_nights / day.total_inventory * 100) if day.total_inventory and day.baseline_room_nights else 0
        
        occupancy_trends.append({
            'date': day.date,
            'occupancy': occupancy,
            'baseline_occupancy': baseline_occupancy
        })
        
        adr_trends.append({
            'date': day.date,
            'adr': day.avg_adr or 0,
            'baseline_adr': day.baseline_adr or 0
        })
    
    return {
        'occupancy_trends': occupancy_trends,
        'adr_trends': adr_trends
    }

def calculate_event_impact_analysis(start_date, end_date, hotel_filter=None):
    """Calculate event impact on hotel performance"""
    # Get events within the date range
    events = Event.query.filter(
        Event.end_date >= start_date,
        Event.start_date <= end_date
    ).all()
    
    event_impacts = []
    
    for event in events:
        # Get hotels in the same city as the event
        city_hotels = Hotel.query.filter_by(city=event.city).all()
        
        event_revenue = 0
        baseline_revenue = 0
        event_room_nights = 0
        baseline_room_nights = 0
        
        for hotel in city_hotels:
            # Get actuals for the event period
            event_actuals = HotelActuals.query.filter(
                HotelActuals.hotel_id == hotel.id,
                HotelActuals.date.between(event.start_date, event.end_date)
            ).all()
            
            for actual in event_actuals:
                if actual.ty_revenue:
                    event_revenue += actual.ty_revenue
                if actual.stly_revenue:
                    baseline_revenue += actual.stly_revenue
                if actual.ty_room_nights:
                    event_room_nights += actual.ty_room_nights
                if actual.stly_room_nights:
                    baseline_room_nights += actual.stly_room_nights
        
        revenue_lift = event_revenue - baseline_revenue if baseline_revenue else 0
        revenue_lift_pct = (revenue_lift / baseline_revenue * 100) if baseline_revenue else 0
        
        event_impacts.append({
            'event': event,
            'hotels_affected': len(city_hotels),
            'event_revenue': event_revenue,
            'baseline_revenue': baseline_revenue,
            'revenue_lift': revenue_lift,
            'revenue_lift_pct': revenue_lift_pct,
            'event_room_nights': event_room_nights,
            'baseline_room_nights': baseline_room_nights
        })
    
    # Sort by revenue lift percentage
    event_impacts.sort(key=lambda x: x['revenue_lift_pct'], reverse=True)
    
    return event_impacts[:10]  # Top 10 most impactful events

def calculate_user_productivity(start_date, end_date):
    """Calculate user productivity and activity metrics"""
    # Get forecast activity by user
    user_forecasts = db.session.query(
        EventForecast.created_by,
        db.func.count(EventForecast.id).label('forecast_count'),
        db.func.count(db.distinct(EventForecast.hotel_id)).label('hotels_worked'),
        db.func.count(db.distinct(EventForecast.event_id)).label('events_worked'),
        db.func.max(EventForecast.created_at).label('last_activity')
    ).filter(
        EventForecast.created_at.between(start_date, datetime.combine(end_date, datetime.max.time()))
    ).group_by(EventForecast.created_by).all()
    
    # Get user task activity - join with User to get username
    user_tasks = db.session.query(
        User.username,
        db.func.count(Task.id).label('tasks_assigned'),
        db.func.count(db.case((Task.status == 'completed', 1))).label('tasks_completed')
    ).join(
        Task, User.id == Task.assigned_to_id
    ).filter(
        Task.created_at.between(start_date, datetime.combine(end_date, datetime.max.time()))
    ).group_by(User.username).all()
    
    productivity_data = []
    
    for forecast_data in user_forecasts:
        username = forecast_data.created_by
        
        # Find corresponding task data
        task_data = next((t for t in user_tasks if t.username == username), None)
        
        productivity_data.append({
            'username': username,
            'forecast_count': forecast_data.forecast_count,
            'hotels_worked': forecast_data.hotels_worked,
            'events_worked': forecast_data.events_worked,
            'last_activity': forecast_data.last_activity,
            'tasks_assigned': task_data.tasks_assigned if task_data else 0,
            'tasks_completed': task_data.tasks_completed if task_data else 0,
            'task_completion_rate': (task_data.tasks_completed / task_data.tasks_assigned * 100) if task_data and task_data.tasks_assigned else 0
        })
    
    # Sort by forecast count
    productivity_data.sort(key=lambda x: x['forecast_count'], reverse=True)
    
    return productivity_data

def calculate_market_comparison(start_date, end_date):
    """Calculate market comparison and benchmarking data"""
    # Get city-level performance
    city_performance = db.session.query(
        Hotel.city,
        db.func.sum(HotelActuals.ty_revenue).label('total_revenue'),
        db.func.sum(HotelActuals.stly_revenue).label('baseline_revenue'),
        db.func.sum(HotelActuals.ty_room_nights).label('total_room_nights'),
        db.func.avg(HotelActuals.ty_adr).label('avg_adr'),
        db.func.count(db.distinct(Hotel.id)).label('hotel_count'),
        db.func.sum(Hotel.inventory).label('total_inventory')
    ).join(
        HotelActuals, Hotel.id == HotelActuals.hotel_id
    ).filter(
        HotelActuals.date.between(start_date, end_date),
        HotelActuals.ty_revenue.isnot(None)
    ).group_by(Hotel.city).all()
    
    market_data = []
    
    for city_data in city_performance:
        days_in_period = (end_date - start_date).days + 1
        available_rooms = city_data.total_inventory * days_in_period if city_data.total_inventory else 1
        occupancy = (city_data.total_room_nights / available_rooms * 100) if available_rooms else 0
        
        revenue_growth = ((city_data.total_revenue - city_data.baseline_revenue) / city_data.baseline_revenue * 100) if city_data.baseline_revenue else 0
        
        market_data.append({
            'city': city_data.city,
            'total_revenue': city_data.total_revenue,
            'revenue_growth': revenue_growth,
            'occupancy': occupancy,
            'avg_adr': city_data.avg_adr or 0,
            'hotel_count': city_data.hotel_count,
            'total_inventory': city_data.total_inventory
        })
    
    # Sort by total revenue
    market_data.sort(key=lambda x: x['total_revenue'], reverse=True)
    
    return market_data

def generate_hotel_analytics(hotel_id):
    """Generate comprehensive analytics for a specific hotel"""
    hotel = Hotel.query.get(hotel_id)
    if not hotel:
        return {}
    
    # Get all available hotel actuals data (no date restriction)
    actuals = HotelActuals.query.filter(
        HotelActuals.hotel_id == hotel_id,
        HotelActuals.ty_revenue.isnot(None)
    ).order_by(HotelActuals.date.desc()).all()
    
    if not actuals:
        return {
            'has_data': False,
            'message': 'No actual performance data available for this hotel'
        }
    
    # Calculate summary metrics
    total_revenue_ty = sum(actual.ty_revenue for actual in actuals if actual.ty_revenue)
    total_revenue_stly = sum(actual.stly_revenue for actual in actuals if actual.stly_revenue)
    total_room_nights_ty = sum(actual.ty_room_nights for actual in actuals if actual.ty_room_nights)
    total_room_nights_stly = sum(actual.stly_room_nights for actual in actuals if actual.stly_room_nights)
    
    # Calculate averages and growth
    avg_adr_ty = total_revenue_ty / total_room_nights_ty if total_room_nights_ty else 0
    avg_adr_stly = total_revenue_stly / total_room_nights_stly if total_room_nights_stly else 0
    
    revenue_growth = ((total_revenue_ty - total_revenue_stly) / total_revenue_stly * 100) if total_revenue_stly else 0
    adr_growth = ((avg_adr_ty - avg_adr_stly) / avg_adr_stly * 100) if avg_adr_stly else 0
    room_nights_growth = ((total_room_nights_ty - total_room_nights_stly) / total_room_nights_stly * 100) if total_room_nights_stly else 0
    
    # Calculate occupancy
    days_with_data = len(actuals)
    available_rooms = hotel.inventory * days_with_data if hotel.inventory else 1
    occupancy_ty = (total_room_nights_ty / available_rooms * 100) if available_rooms else 0
    occupancy_stly = (total_room_nights_stly / available_rooms * 100) if available_rooms else 0
    occupancy_growth = occupancy_ty - occupancy_stly
    
    # Prepare table data
    table_data = []
    for actual in actuals[:30]:  # Last 30 days for table
        # Calculate current year occupancy
        ty_occupancy = (actual.ty_room_nights / hotel.inventory * 100) if actual.ty_room_nights and hotel.inventory else 0
        stly_occupancy = (actual.stly_room_nights / hotel.inventory * 100) if actual.stly_room_nights and hotel.inventory else 0
        
        # Calculate growth rates
        revenue_growth_daily = ((actual.ty_revenue - actual.stly_revenue) / actual.stly_revenue * 100) if actual.ty_revenue and actual.stly_revenue else 0
        adr_growth_daily = ((actual.ty_adr - actual.stly_adr) / actual.stly_adr * 100) if actual.ty_adr and actual.stly_adr else 0
        
        table_data.append({
            'date': actual.date,
            'ty_revenue': actual.ty_revenue or 0,
            'stly_revenue': actual.stly_revenue or 0,
            'revenue_growth': revenue_growth_daily,
            'ty_adr': actual.ty_adr or 0,
            'stly_adr': actual.stly_adr or 0,
            'adr_growth': adr_growth_daily,
            'ty_room_nights': actual.ty_room_nights or 0,
            'stly_room_nights': actual.stly_room_nights or 0,
            'ty_occupancy': ty_occupancy,
            'stly_occupancy': stly_occupancy,
            'occupancy_change': ty_occupancy - stly_occupancy
        })
    
    # Prepare chart data
    chart_data = {
        'dates': [actual.date.strftime('%m/%d') for actual in reversed(actuals)],
        'revenue_ty': [actual.ty_revenue or 0 for actual in reversed(actuals)],
        'revenue_stly': [actual.stly_revenue or 0 for actual in reversed(actuals)],
        'adr_ty': [actual.ty_adr or 0 for actual in reversed(actuals)],
        'adr_stly': [actual.stly_adr or 0 for actual in reversed(actuals)],
        'occupancy_ty': [(actual.ty_room_nights / hotel.inventory * 100) if actual.ty_room_nights and hotel.inventory else 0 for actual in reversed(actuals)],
        'occupancy_stly': [(actual.stly_room_nights / hotel.inventory * 100) if actual.stly_room_nights and hotel.inventory else 0 for actual in reversed(actuals)]
    }
    
    return {
        'has_data': True,
        'summary': {
            'total_revenue_ty': total_revenue_ty,
            'total_revenue_stly': total_revenue_stly,
            'revenue_growth': revenue_growth,
            'avg_adr_ty': avg_adr_ty,
            'avg_adr_stly': avg_adr_stly,
            'adr_growth': adr_growth,
            'total_room_nights_ty': total_room_nights_ty,
            'total_room_nights_stly': total_room_nights_stly,
            'room_nights_growth': room_nights_growth,
            'occupancy_ty': occupancy_ty,
            'occupancy_stly': occupancy_stly,
            'occupancy_growth': occupancy_growth,
            'days_analyzed': days_with_data
        },
        'table_data': table_data,
        'chart_data': chart_data
    }

# Comment Routes
@app.route('/add_comment', methods=['POST'])
@login_required
def add_comment():
    """Add a comment to a hotel or event"""
    content = request.form.get('content', '').strip()
    hotel_id = request.form.get('hotel_id')
    event_id = request.form.get('event_id')
    parent_id = request.form.get('parent_id')
    
    if not content:
        flash('Comment cannot be empty.', 'error')
        return redirect(request.referrer or url_for('index'))
    
    comment = Comment(
        content=content,
        user_id=current_user.id,
        hotel_id=int(hotel_id) if hotel_id else None,
        event_id=int(event_id) if event_id else None,
        parent_id=int(parent_id) if parent_id else None
    )
    
    try:
        db.session.add(comment)
        db.session.commit()
        flash('Comment added successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error adding comment. Please try again.', 'error')
    
    return redirect(request.referrer or url_for('index'))

@app.route('/delete_comment/<int:comment_id>', methods=['POST'])
@login_required
def delete_comment(comment_id):
    """Delete a comment (only the author can delete)"""
    comment = Comment.query.get_or_404(comment_id)
    
    if comment.user_id != current_user.id:
        flash('You can only delete your own comments.', 'error')
        return redirect(request.referrer or url_for('index'))
    
    try:
        # Delete all replies first
        Comment.query.filter_by(parent_id=comment_id).delete()
        db.session.delete(comment)
        db.session.commit()
        flash('Comment deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting comment. Please try again.', 'error')
    
    return redirect(request.referrer or url_for('index'))

@app.route('/user-activity')
def user_activity():
    # Get filter parameters
    user_filter = request.args.get('user_filter', '')
    action_filter = request.args.get('action_filter', '')
    date_filter = request.args.get('date_filter', '')
    
    # Get all unique users from different tables (HotelActuals doesn't have created_by)
    forecast_users = db.session.query(EventForecast.created_by).distinct().all()
    monthly_forecast_users = db.session.query(MonthlyForecast.created_by).distinct().all()
    
    all_users = set()
    for user_tuple in forecast_users + monthly_forecast_users:
        if user_tuple[0]:
            all_users.add(user_tuple[0])
    all_users = sorted(list(all_users))
    
    # Build activity data from forecasts and actuals
    activities = []
    
    # Get forecast activities
    forecast_query = db.session.query(
        EventForecast.created_by,
        EventForecast.created_at,
        Hotel.hotel_name,
        Event.event_name
    ).join(Hotel).join(Event)
    
    if user_filter:
        forecast_query = forecast_query.filter(EventForecast.created_by == user_filter)
    
    if date_filter:
        today = datetime.now().date()
        if date_filter == 'today':
            forecast_query = forecast_query.filter(EventForecast.created_at >= today)
        elif date_filter == 'week':
            week_ago = today - timedelta(days=7)
            forecast_query = forecast_query.filter(EventForecast.created_at >= week_ago)
        elif date_filter == 'month':
            month_ago = today - timedelta(days=30)
            forecast_query = forecast_query.filter(EventForecast.created_at >= month_ago)
    
    if not action_filter or action_filter == 'forecast':
        for user, timestamp, hotel_name, event_name in forecast_query.order_by(desc(EventForecast.created_at)).limit(50):
            if user:
                activities.append({
                    'user': user,
                    'action': f'Created forecast for {hotel_name} - {event_name}',
                    'type': 'Forecast',
                    'timestamp': timestamp
                })
    
    # Note: HotelActuals doesn't have created_by field, so we skip actuals activities for now
    # This can be added later when the model is updated with user tracking
    
    # Add monthly forecast activities
    monthly_query = db.session.query(
        MonthlyForecast.created_by,
        MonthlyForecast.created_at,
        Hotel.hotel_name,
        MonthlyForecast.forecast_date
    ).join(Hotel)
    
    if user_filter:
        monthly_query = monthly_query.filter(MonthlyForecast.created_by == user_filter)
    
    if date_filter:
        today = datetime.now().date()
        if date_filter == 'today':
            monthly_query = monthly_query.filter(MonthlyForecast.created_at >= today)
        elif date_filter == 'week':
            week_ago = today - timedelta(days=7)
            monthly_query = monthly_query.filter(MonthlyForecast.created_at >= week_ago)
        elif date_filter == 'month':
            month_ago = today - timedelta(days=30)
            monthly_query = monthly_query.filter(MonthlyForecast.created_at >= month_ago)
    
    if not action_filter or action_filter == 'monthly':
        for user, timestamp, hotel_name, forecast_date in monthly_query.order_by(desc(MonthlyForecast.created_at)).limit(50):
            if user:
                activities.append({
                    'user': user,
                    'action': f'Created monthly forecast for {hotel_name} on {forecast_date.strftime("%Y-%m-%d")}',
                    'type': 'Monthly Forecast',
                    'timestamp': timestamp
                })
    
    # Sort activities by timestamp
    activities.sort(key=lambda x: x['timestamp'], reverse=True)
    activities = activities[:100]  # Limit to 100 most recent
    
    # Get statistics
    total_users = len(all_users)
    total_forecasts = EventForecast.query.count()
    total_monthly_forecasts = MonthlyForecast.query.count()
    total_hotels = Hotel.query.count()
    total_events = Event.query.count()
    
    return render_template('user_activity.html',
                         activities=activities,
                         all_users=all_users,
                         total_users=total_users,
                         total_forecasts=total_forecasts,
                         total_monthly_forecasts=total_monthly_forecasts,
                         total_hotels=total_hotels,
                         total_events=total_events,
                         user_filter=user_filter,
                         action_filter=action_filter,
                         date_filter=date_filter)

@app.route('/upload-actuals', methods=['GET', 'POST'])
def upload_actuals():
    """Handle hotel actuals data upload and processing"""
    
    if request.method == 'GET':
        # Show upload form
        return render_template('upload_actuals.html')
    
    if 'actuals_file' not in request.files:
        flash('No file selected. Please choose a file to upload.', 'error')
        return redirect(request.url)
    
    file = request.files['actuals_file']
    if file.filename == '':
        flash('No file selected. Please choose a file to upload.', 'error')
        return redirect(request.url)
    
    if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
        flash('Invalid file format. Please upload a CSV or Excel file.', 'error')
        return redirect(request.url)
    
    try:
        # Read uploaded file
        if file.filename.lower().endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        # Validate minimum number of columns (use column order instead of exact names)
        expected_column_count = 11  # Date, Hotel Code, Revenue TY, Room Nights TY, ADR TY, STLY Revenue, STLY Room Nights, STLY ADR, Result LY Revenue, Result LY Room Nights, Result LY ADR
        
        if len(df.columns) < expected_column_count:
            flash(f'File must have at least {expected_column_count} columns. Found {len(df.columns)} columns. Please check the template format.', 'error')
            return redirect(request.url)
        
        # Use column positions instead of names for better flexibility
        # Expected column order: Date, Hotel Code, Revenue TY, Room Nights TY, ADR TY, STLY Revenue, STLY Room Nights, STLY ADR, Result LY Revenue, Result LY Room Nights, Result LY ADR
        column_mapping = {
            'date': 0,
            'hotel_code': 1, 
            'revenue_ty': 2,
            'room_nights_ty': 3,
            'adr_ty': 4,
            'stly_revenue': 5,
            'stly_room_nights': 6,
            'stly_adr': 7,
            'resly_revenue': 8,
            'resly_room_nights': 9,
            'resly_adr': 10
        }
        
        # Process and validate data
        successful_updates = 0
        skipped_rows = 0
        error_messages = []
        
        user_name = session.get('user_name', 'Anonymous')
        
        for index, row in df.iterrows():
            try:
                # Parse date using column position
                row_date = pd.to_datetime(row.iloc[column_mapping['date']]).date()
                
                # Find hotel by code using column position
                hotel_code = str(row.iloc[column_mapping['hotel_code']]).strip()
                hotel = Hotel.query.filter_by(hotel_code=hotel_code).first()
                if not hotel:
                    skipped_rows += 1
                    error_messages.append(f"Row {index + 2}: Hotel code '{hotel_code}' not found")
                    continue
                
                # Check if row has any actual data to process using column positions
                has_data = any([
                    pd.notna(row.iloc[column_mapping['revenue_ty']]) and row.iloc[column_mapping['revenue_ty']] != '',
                    pd.notna(row.iloc[column_mapping['room_nights_ty']]) and row.iloc[column_mapping['room_nights_ty']] != '',
                    pd.notna(row.iloc[column_mapping['adr_ty']]) and row.iloc[column_mapping['adr_ty']] != '',
                    pd.notna(row.iloc[column_mapping['stly_revenue']]) and row.iloc[column_mapping['stly_revenue']] != '',
                    pd.notna(row.iloc[column_mapping['stly_room_nights']]) and row.iloc[column_mapping['stly_room_nights']] != '',
                    pd.notna(row.iloc[column_mapping['stly_adr']]) and row.iloc[column_mapping['stly_adr']] != '',
                    pd.notna(row.iloc[column_mapping['resly_revenue']]) and row.iloc[column_mapping['resly_revenue']] != '',
                    pd.notna(row.iloc[column_mapping['resly_room_nights']]) and row.iloc[column_mapping['resly_room_nights']] != '',
                    pd.notna(row.iloc[column_mapping['resly_adr']]) and row.iloc[column_mapping['resly_adr']] != ''
                ])
                
                if not has_data:
                    continue  # Skip empty rows
                
                # Find or create HotelActuals record
                actual = HotelActuals.query.filter_by(
                    date=row_date,
                    hotel_id=hotel.id
                ).first()
                
                if not actual:
                    actual = HotelActuals(date=row_date, hotel_id=hotel.id, created_by=user_name)
                    db.session.add(actual)
                else:
                    actual.updated_at = datetime.utcnow()
                    actual.created_by = user_name
                
                # Update fields with validation
                def safe_float(value):
                    try:
                        return float(value) if pd.notna(value) and value != '' else None
                    except (ValueError, TypeError):
                        return None
                
                def safe_int(value):
                    try:
                        return int(value) if pd.notna(value) and value != '' else None
                    except (ValueError, TypeError):
                        return None
                
                # Update TY (This Year) data using column positions
                actual.ty_revenue = safe_float(row.iloc[column_mapping['revenue_ty']])
                actual.ty_room_nights = safe_int(row.iloc[column_mapping['room_nights_ty']])
                actual.ty_adr = safe_float(row.iloc[column_mapping['adr_ty']])
                
                # Update STLY (Same Time Last Year) data using column positions
                actual.stly_revenue = safe_float(row.iloc[column_mapping['stly_revenue']])
                actual.stly_room_nights = safe_int(row.iloc[column_mapping['stly_room_nights']])
                actual.stly_adr = safe_float(row.iloc[column_mapping['stly_adr']])
                
                # Update RESLY (Results Last Year) data using column positions
                actual.resly_revenue = safe_float(row.iloc[column_mapping['resly_revenue']])
                actual.resly_room_nights = safe_int(row.iloc[column_mapping['resly_room_nights']])
                actual.resly_adr = safe_float(row.iloc[column_mapping['resly_adr']])
                
                # Calculate occupancy if room nights and inventory are available
                actual.calculate_occupancy()
                
                successful_updates += 1
                
            except Exception as e:
                skipped_rows += 1
                error_messages.append(f"Row {index + 2}: {str(e)}")
                continue
        
        # Commit all changes
        db.session.commit()
        
        # Create success message
        message_parts = [f'Successfully processed {successful_updates} records']
        if skipped_rows > 0:
            message_parts.append(f'{skipped_rows} rows were skipped')
        
        flash('. '.join(message_parts) + '.', 'success')
        
        # Show detailed error messages if any
        if error_messages and len(error_messages) <= 10:  # Limit error messages
            for error_msg in error_messages[:10]:
                flash(error_msg, 'warning')
        elif len(error_messages) > 10:
            flash(f'Plus {len(error_messages) - 10} additional errors. Please check your data format.', 'warning')
        
        return redirect(url_for('actuals_dashboard'))
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'error')
        return redirect(request.url)

@app.route('/actuals-dashboard')
@login_required
def actuals_dashboard():
    """Display uploaded hotel actuals data with comprehensive analysis"""
    
    # Get filter parameters
    hotel_filter = request.args.get('hotel', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    event_filter = request.args.get('event', '')
    
    # Build query
    query = db.session.query(HotelActuals, Hotel).join(Hotel, HotelActuals.hotel_id == Hotel.id)
    
    if hotel_filter:
        query = query.filter(or_(
            Hotel.hotel_code.ilike(f'%{hotel_filter}%'),
            Hotel.hotel_name.ilike(f'%{hotel_filter}%')
        ))
    
    # Event filtering logic
    selected_event = None
    if event_filter:
        try:
            event_id = int(event_filter)
            selected_event = Event.query.get(event_id)
            if selected_event:
                # Filter dates to the event period
                query = query.filter(
                    HotelActuals.date >= selected_event.start_date,
                    HotelActuals.date <= selected_event.end_date
                )
        except (ValueError, TypeError):
            pass
    
    if date_from and not event_filter:
        try:
            date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(HotelActuals.date >= date_from_parsed)
        except ValueError:
            pass
    
    if date_to and not event_filter:
        try:
            date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(HotelActuals.date <= date_to_parsed)
        except ValueError:
            pass
    
    # Get results ordered by date and hotel
    actuals_data = query.order_by(HotelActuals.date.desc(), Hotel.hotel_code).limit(500).all()
    
    # Get all hotels and events for filter dropdowns
    all_hotels = Hotel.query.order_by(Hotel.hotel_code).all()
    all_events = Event.query.order_by(Event.start_date.desc()).all()
    
    # Calculate summary statistics
    total_records = query.count()
    
    # Separate data into categories for analysis
    actual_table_data = []
    baseline_table_data = []
    impact_table_data = []
    
    # Data for charts
    chart_dates = []
    actual_revenue_data = []
    baseline_revenue_data = []
    impact_revenue_data = []
    actual_adr_data = []
    baseline_adr_data = []
    actual_occupancy_data = []
    baseline_occupancy_data = []
    
    # Get forecasts data if event is selected
    forecast_data = {}
    if selected_event and hotel_filter:
        # Find hotel by filter
        filtered_hotel = Hotel.query.filter(or_(
            Hotel.hotel_code.ilike(f'%{hotel_filter}%'),
            Hotel.hotel_name.ilike(f'%{hotel_filter}%')
        )).first()
        
        if filtered_hotel:
            # Get forecast data for this event and hotel
            event_forecasts = EventForecast.query.filter_by(
                event_id=selected_event.id,
                hotel_id=filtered_hotel.id
            ).all()
            
            for forecast in event_forecasts:
                # Calculate the actual date based on event start date and day number
                forecast_date = selected_event.start_date + timedelta(days=forecast.day_number - 1)
                forecast_data[forecast_date] = {
                    'revenue': forecast.revenue,
                    'adr': forecast.adr,
                    'occupancy': forecast.occupancy,
                    'created_by': forecast.created_by
                }
    
    for actual, hotel in actuals_data:
        # Chart data collection
        if actual.date not in chart_dates:
            chart_dates.append(actual.date)
            
        # Actual Performance Table Data
        if actual.ty_revenue or actual.ty_room_nights or actual.ty_adr:
            # Calculate occupancy from room nights and inventory
            occupancy_ty = (actual.ty_room_nights / hotel.inventory * 100) if actual.ty_room_nights and hotel.inventory else None
            actual_table_data.append({
                'date': actual.date,
                'hotel': hotel,
                'revenue': actual.ty_revenue,
                'room_nights': actual.ty_room_nights,
                'adr': actual.ty_adr,
                'occupancy': occupancy_ty,
                'forecast': forecast_data.get(actual.date, {}) if forecast_data else {}
            })
            
            # Collect chart data
            if actual.ty_revenue:
                actual_revenue_data.append({'x': actual.date, 'y': actual.ty_revenue, 'hotel': hotel.hotel_code})
            if actual.ty_adr:
                actual_adr_data.append({'x': actual.date, 'y': actual.ty_adr, 'hotel': hotel.hotel_code})
            if occupancy_ty:
                actual_occupancy_data.append({'x': actual.date, 'y': occupancy_ty, 'hotel': hotel.hotel_code})
        
        # Baseline (STLY) Performance Table Data
        if actual.stly_revenue or actual.stly_room_nights or actual.stly_adr:
            baseline_occupancy = (actual.stly_room_nights / hotel.inventory * 100) if actual.stly_room_nights and hotel.inventory else None
            baseline_table_data.append({
                'date': actual.date,
                'hotel': hotel,
                'revenue': actual.stly_revenue,
                'room_nights': actual.stly_room_nights,
                'adr': actual.stly_adr,
                'occupancy': baseline_occupancy
            })
            
            # Collect baseline chart data
            if actual.stly_revenue:
                baseline_revenue_data.append({'x': actual.date, 'y': actual.stly_revenue, 'hotel': hotel.hotel_code})
            if actual.stly_adr:
                baseline_adr_data.append({'x': actual.date, 'y': actual.stly_adr, 'hotel': hotel.hotel_code})
            if baseline_occupancy:
                baseline_occupancy_data.append({'x': actual.date, 'y': baseline_occupancy, 'hotel': hotel.hotel_code})
        
        # Impact Analysis Table Data
        if (actual.ty_revenue and actual.stly_revenue) or (actual.ty_room_nights and actual.stly_room_nights):
            revenue_impact = actual.ty_revenue - actual.stly_revenue if actual.ty_revenue and actual.stly_revenue else None
            adr_impact = actual.ty_adr - actual.stly_adr if actual.ty_adr and actual.stly_adr else None
            room_nights_impact = actual.ty_room_nights - actual.stly_room_nights if actual.ty_room_nights and actual.stly_room_nights else None
            
            # Calculate occupancy impact
            occupancy_ty = (actual.ty_room_nights / hotel.inventory * 100) if actual.ty_room_nights and hotel.inventory else None
            occupancy_stly = (actual.stly_room_nights / hotel.inventory * 100) if actual.stly_room_nights and hotel.inventory else None
            occupancy_impact = occupancy_ty - occupancy_stly if occupancy_ty is not None and occupancy_stly is not None else None
            
            impact_table_data.append({
                'date': actual.date,
                'hotel': hotel,
                'revenue_impact': revenue_impact,
                'adr_impact': adr_impact,
                'room_nights_impact': room_nights_impact,
                'occupancy_impact': occupancy_impact
            })
            
            # Collect impact chart data
            if revenue_impact:
                impact_revenue_data.append({'x': actual.date, 'y': revenue_impact, 'hotel': hotel.hotel_code})
    
    # Create interactive charts
    charts = create_performance_charts(
        actual_revenue_data, baseline_revenue_data, impact_revenue_data,
        actual_adr_data, baseline_adr_data,
        actual_occupancy_data, baseline_occupancy_data
    )
    
    return render_template('actuals_dashboard.html',
                         actuals_data=actuals_data,
                         actual_table_data=actual_table_data,
                         baseline_table_data=baseline_table_data,
                         impact_table_data=impact_table_data,
                         actual_revenue_data=[(item[0].date, item[0].ty_revenue or 0, item[0].stly_revenue or 0) for item in actuals_data],
                         actual_adr_data=[(item[0].date, item[0].ty_adr or 0, item[0].stly_adr or 0) for item in actuals_data],
                         actual_occupancy_data=[(item[0].date, 
                                               (item[0].ty_room_nights / item[1].inventory * 100) if item[1].inventory and item[0].ty_room_nights else 0,
                                               (item[0].stly_room_nights / item[1].inventory * 100) if item[1].inventory and item[0].stly_room_nights else 0) 
                                              for item in actuals_data],
                         all_hotels=all_hotels,
                         all_events=all_events,
                         selected_event=selected_event,
                         total_records=total_records,
                         charts=charts,
                         current_filters={
                             'hotel': hotel_filter,
                             'date_from': date_from,
                             'date_to': date_to,
                             'event': event_filter
                         })

@app.route('/chat')
def chat():
    """AI Chat interface"""
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return render_template('chat.html')

@app.route('/chat/message', methods=['POST'])
def chat_message():
    """Handle chat messages with conversation saving"""
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'error': 'Authentication required'})
    
    try:
        import time
        start_time = time.time()
        
        data = request.get_json()
        message = data.get('message', '').strip()
        conversation_id = data.get('conversation_id')
        
        if not message:
            return jsonify({'success': False, 'error': 'Empty message'})
        
        # Get or create conversation
        if conversation_id:
            conversation = ChatConversation.query.filter_by(
                id=conversation_id, 
                user_id=current_user.id
            ).first()
        else:
            conversation = None
            
        if not conversation:
            # Create new conversation with title from first message
            title = message[:50] + "..." if len(message) > 50 else message
            conversation = ChatConversation(
                title=title,
                user_id=current_user.id
            )
            db.session.add(conversation)
            db.session.flush()  # Get the ID
        
        # Process message with AI assistant
        response = chat_assistant.process_message(message)
        
        # Save message and response
        chat_message = ChatMessage(
            conversation_id=conversation.id,
            message_text=message,
            response_text=response,
            is_user_message=True,
            processing_time=time.time() - start_time
        )
        db.session.add(chat_message)
        
        # Update conversation timestamp
        conversation.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'response': response,
            'conversation_id': conversation.id
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Chat error: {e}")
        return jsonify({
            'success': False, 
            'error': 'I encountered an error processing your request.'
        })

@app.route('/chat/conversation', methods=['POST'])
def create_conversation():
    """Create or get current conversation"""
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'error': 'Authentication required'})
    
    try:
        # Get the most recent conversation or create new one
        conversation = ChatConversation.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).order_by(desc(ChatConversation.updated_at)).first()
        
        if not conversation:
            conversation = ChatConversation(
                title="New Chat",
                user_id=current_user.id
            )
            db.session.add(conversation)
            db.session.commit()
        
        return jsonify({
            'success': True,
            'conversation_id': conversation.id
        })
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Conversation creation error: {e}")
        return jsonify({'success': False, 'error': 'Failed to create conversation'})

@app.route('/chat/conversations')
def get_conversations():
    """Get user's chat conversations"""
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'error': 'Authentication required'})
    
    try:
        conversations = ChatConversation.query.filter_by(
            user_id=current_user.id,
            is_active=True
        ).order_by(desc(ChatConversation.updated_at)).limit(20).all()
        
        conversations_data = []
        for conv in conversations:
            # Get first message as preview
            first_message = ChatMessage.query.filter_by(
                conversation_id=conv.id
            ).order_by(ChatMessage.created_at).first()
            
            preview = first_message.message_text[:100] + "..." if first_message and len(first_message.message_text) > 100 else (first_message.message_text if first_message else "")
            
            conversations_data.append({
                'id': conv.id,
                'title': conv.title,
                'created_at': conv.created_at.isoformat(),
                'updated_at': conv.updated_at.isoformat(),
                'preview': preview
            })
        
        return jsonify({
            'success': True,
            'conversations': conversations_data
        })
        
    except Exception as e:
        logging.error(f"Get conversations error: {e}")
        return jsonify({'success': False, 'error': 'Failed to load conversations'})

@app.route('/chat/conversation/<int:conversation_id>/messages')
def get_conversation_messages(conversation_id):
    """Get messages for a specific conversation"""
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'error': 'Authentication required'})
    
    try:
        # Verify user owns the conversation
        conversation = ChatConversation.query.filter_by(
            id=conversation_id,
            user_id=current_user.id
        ).first()
        
        if not conversation:
            return jsonify({'success': False, 'error': 'Conversation not found'})
        
        messages = ChatMessage.query.filter_by(
            conversation_id=conversation_id
        ).order_by(ChatMessage.created_at).all()
        
        messages_data = []
        for msg in messages:
            messages_data.append({
                'id': msg.id,
                'message_text': msg.message_text,
                'response_text': msg.response_text,
                'is_user_message': msg.is_user_message,
                'created_at': msg.created_at.isoformat()
            })
        
        return jsonify({
            'success': True,
            'messages': messages_data
        })
        
    except Exception as e:
        logging.error(f"Get messages error: {e}")
        return jsonify({'success': False, 'error': 'Failed to load messages'})

# Event Finder Routes
@app.route('/event-finder')
@login_required
def event_finder():
    """Event finder main page"""
    # Get user's recent searches
    recent_searches = EventSearch.query.filter_by(user_id=current_user.id).order_by(EventSearch.created_at.desc()).limit(10).all()
    return render_template('events/event_finder.html', recent_searches=recent_searches)

@app.route('/event-finder/search', methods=['POST'])
@login_required
def event_finder_search():
    """Perform event search"""
    try:
        location = request.form.get('location', '').strip()
        event_types = request.form.getlist('event_types')
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        search_name = request.form.get('search_name', '').strip()
        
        # Validate input
        if not location:
            flash('Please enter a location.', 'error')
            return redirect(url_for('event_finder'))
        
        if not start_date_str or not end_date_str:
            flash('Please select both start and end dates.', 'error')
            return redirect(url_for('event_finder'))
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        if start_date > end_date:
            flash('Start date cannot be after end date.', 'error')
            return redirect(url_for('event_finder'))
        
        # Initialize event finder service
        event_service = EventFinderService()
        
        # Perform search
        search = event_service.search_events(
            location=location,
            event_types=event_types,
            start_date=start_date,
            end_date=end_date,
            user_id=current_user.id,
            search_name=search_name
        )
        
        flash(f'Found {search.results_count} events for your search!', 'success')
        return redirect(url_for('event_search_results', search_id=search.id))
        
    except Exception as e:
        flash(f'Error performing search: {str(e)}', 'error')
        return redirect(url_for('event_finder'))

@app.route('/event-finder/results/<int:search_id>')
@login_required
def event_search_results(search_id):
    """Display event search results"""
    search = EventSearch.query.get_or_404(search_id)
    
    # Verify user owns this search
    if search.user_id != current_user.id:
        flash('You can only view your own searches.', 'error')
        return redirect(url_for('event_finder'))
    
    # Parse event types for template
    try:
        search_event_types = json.loads(search.event_types) if search.event_types else []
    except:
        search_event_types = []
    
    # Get events with pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    events = ExternalEvent.query.filter_by(search_id=search_id).order_by(ExternalEvent.event_date.asc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('events/search_results.html', search=search, events=events, search_event_types=search_event_types)

@app.route('/event-finder/export/<int:search_id>')
@login_required
def export_event_search(search_id):
    """Export event search results to Excel"""
    try:
        event_service = EventFinderService()
        excel_buffer, filename = event_service.export_to_excel(search_id, current_user.id)
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error exporting data: {str(e)}', 'error')
        return redirect(url_for('event_search_results', search_id=search_id))

@app.route('/event-finder/favorite/<int:event_id>', methods=['POST'])
@login_required
def toggle_event_favorite(event_id):
    """Toggle favorite status for an event"""
    try:
        event_service = EventFinderService()
        success = event_service.favorite_event(event_id, current_user.id)
        
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Access denied'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/event-finder/note/<int:event_id>', methods=['POST'])
@login_required
def update_event_note(event_id):
    """Update note for an event"""
    try:
        note = request.json.get('note', '').strip()
        event_service = EventFinderService()
        success = event_service.add_event_note(event_id, current_user.id, note)
        
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Access denied'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/event-finder/history')
@login_required
def event_search_history():
    """Display user's event search history"""
    event_service = EventFinderService()
    searches = event_service.get_search_history(current_user.id)
    
    # Parse event types for each search
    for search in searches:
        try:
            search.parsed_event_types = json.loads(search.event_types) if search.event_types else []
        except:
            search.parsed_event_types = []
    
    return render_template('events/search_history.html', searches=searches)

@app.route('/event-finder/import-event/<int:event_id>', methods=['POST'])
@login_required
def import_external_event(event_id):
    """Import a single external event to the main events database"""
    try:
        external_event = ExternalEvent.query.get_or_404(event_id)
        
        # Check if user owns this search
        if external_event.search.user_id != current_user.id:
            return jsonify({'success': False, 'error': 'Access denied'})
        
        # Check if event already exists
        existing_event = Event.query.filter_by(
            event_name=external_event.event_name,
            city=external_event.city,
            start_date=external_event.event_date
        ).first()
        
        if existing_event:
            return jsonify({'success': False, 'error': 'Event already exists in database'})
        
        # Create new Event from ExternalEvent
        new_event = Event(
            event_name=external_event.event_name,
            start_date=external_event.event_date,
            end_date=external_event.end_date or external_event.event_date,
            city=external_event.city,
            website_url=external_event.source_url,
            image_url=None,  # Could be added later
            description=external_event.description
        )
        
        db.session.add(new_event)
        db.session.flush()  # Get the new event ID
        
        # Mark external event as imported
        external_event.is_imported = True
        external_event.imported_event_id = new_event.id
        
        db.session.commit()
        
        return jsonify({'success': True, 'event_id': new_event.id})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/event-finder/import-all/<int:search_id>', methods=['POST'])
@login_required
def import_all_events(search_id):
    """Import all external events from a search to the main events database"""
    try:
        search = EventSearch.query.get_or_404(search_id)
        
        # Check if user owns this search
        if search.user_id != current_user.id:
            return jsonify({'success': False, 'error': 'Access denied'})
        
        external_events = ExternalEvent.query.filter_by(search_id=search_id).all()
        imported_count = 0
        
        for external_event in external_events:
            # Check if event already exists
            existing_event = Event.query.filter_by(
                event_name=external_event.event_name,
                city=external_event.city,
                start_date=external_event.event_date
            ).first()
            
            if not existing_event:
                # Create new Event from ExternalEvent
                new_event = Event(
                    event_name=external_event.event_name,
                    start_date=external_event.event_date,
                    end_date=external_event.end_date or external_event.event_date,
                    city=external_event.city,
                    website_url=external_event.source_url,
                    image_url=None,
                    description=external_event.description
                )
                
                db.session.add(new_event)
                db.session.flush()  # Get the new event ID
                
                # Mark external event as imported
                external_event.is_imported = True
                external_event.imported_event_id = new_event.id
                
                imported_count += 1
        
        db.session.commit()
        
        return jsonify({'success': True, 'imported_count': imported_count})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/event-finder/rerun/<int:search_id>')
@login_required
def rerun_event_search(search_id):
    """Re-run a previous event search"""
    try:
        original_search = EventSearch.query.get_or_404(search_id)
        
        # Verify user owns this search
        if original_search.user_id != current_user.id:
            flash('You can only rerun your own searches.', 'error')
            return redirect(url_for('event_finder'))
        
        # Parse event types
        event_types = json.loads(original_search.event_types) if original_search.event_types else []
        
        # Initialize event finder service
        event_service = EventFinderService()
        
        # Perform new search with same parameters
        new_search = event_service.search_events(
            location=original_search.location,
            event_types=event_types,
            start_date=datetime.combine(original_search.start_date, datetime.min.time()),
            end_date=datetime.combine(original_search.end_date, datetime.min.time()),
            user_id=current_user.id,
            search_name=f"{original_search.search_name} (Rerun)" if original_search.search_name else None
        )
        
        flash(f'Search rerun complete! Found {new_search.results_count} events.', 'success')
        return redirect(url_for('event_search_results', search_id=new_search.id))
        
    except Exception as e:
        flash(f'Error rerunning search: {str(e)}', 'error')
        return redirect(url_for('event_search_history'))
